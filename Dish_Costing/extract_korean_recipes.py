import json
import openpyxl

wb = openpyxl.load_workbook(
    r'C:\Users\harsh\Desktop\Cloud Kitchen\Brands Home\Korean Food\Korean Recipes.xlsx',
    data_only=True
)

def extract_sheet(ws, sheet_type):
    """Extract recipes from a sheet.
    Columns:
      B(2)=Dish Name, C(3)=Ingredient, D(4)=EP Qty, E(5)=UOM,
      F(6)=Wastage%, J(10)=is_semi_finished(No/Yes?), K(11)=Recipe Code,
      L(12)=Ingredient Code, M(13)=Quantity(EP), N(14)=UOM, O(15)=Wastage%
    """
    recipes = {}
    current_dish = None

    # Find header row first, then start from next row
    start_row = 3
    for r in range(1, 6):
        if ws.cell(row=r, column=3).value == 'Ingredient':
            start_row = r + 1
            break

    for row_num in range(start_row, ws.max_row + 1):
        dish_name = ws.cell(row=row_num, column=2).value
        ingredient = ws.cell(row=row_num, column=3).value
        ep_qty = ws.cell(row=row_num, column=4).value
        uom = ws.cell(row=row_num, column=5).value
        wastage = ws.cell(row=row_num, column=6).value
        is_sf = ws.cell(row=row_num, column=10).value  # "No" or similar
        recipe_code = ws.cell(row=row_num, column=11).value
        ingredient_code = ws.cell(row=row_num, column=12).value
        qty_ep = ws.cell(row=row_num, column=13).value
        uom2 = ws.cell(row=row_num, column=14).value
        wastage2 = ws.cell(row=row_num, column=15).value

        if ingredient is None and dish_name is None:
            continue

        if dish_name is not None:
            current_dish = str(dish_name).strip()

        if current_dish is None or ingredient is None:
            continue

        ingredient = str(ingredient).strip()

        # Use primary columns D-F, fallback to M-O if available
        final_qty = ep_qty if ep_qty is not None else qty_ep
        final_uom = str(uom).strip() if uom else (str(uom2).strip() if uom2 else '')
        final_wastage = wastage if wastage is not None else wastage2
        if final_wastage is None:
            final_wastage = 0

        # Determine if ingredient is a semi-finished recipe reference
        is_semi_finished_ref = False
        if ingredient_code and str(ingredient_code).startswith('KOR'):
            is_semi_finished_ref = True
        if ingredient.endswith('- SF'):
            is_semi_finished_ref = True

        ing_entry = {
            'ingredient': ingredient,
            'ingredient_code': str(ingredient_code) if ingredient_code else None,
            'ep_qty': float(final_qty) if final_qty is not None else 0,
            'uom': final_uom,
            'wastage_pct': float(final_wastage) if final_wastage else 0,
            'is_semi_finished_ref': is_semi_finished_ref
        }

        if current_dish not in recipes:
            recipes[current_dish] = {
                'dish_name': current_dish,
                'recipe_code': str(recipe_code) if recipe_code else None,
                'type': sheet_type,
                'ingredients': []
            }
        elif recipe_code and recipes[current_dish]['recipe_code'] is None:
            recipes[current_dish]['recipe_code'] = str(recipe_code)

        recipes[current_dish]['ingredients'].append(ing_entry)

    return list(recipes.values())


# Extract both sheets
semi_finished = extract_sheet(wb['Semi Finished'], 'semi_finished')
finished = extract_sheet(wb['Finished1'], 'finished')

# Build output
output = {
    'title': 'Korean Food Recipes - Cloud Kitchen',
    'source_file': 'Korean Recipes.xlsx',
    'summary': {
        'total_semi_finished_recipes': len(semi_finished),
        'total_finished_recipes': len(finished),
        'total_recipes': len(semi_finished) + len(finished)
    },
    'semi_finished_recipes': semi_finished,
    'finished_recipes': finished
}

# Add ingredient counts
for r in output['semi_finished_recipes']:
    r['ingredient_count'] = len(r['ingredients'])
for r in output['finished_recipes']:
    r['ingredient_count'] = len(r['ingredients'])

out_path = r'E:\Cloud Kitchen\AI Teams\korean_recipes.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False, default=str)

print(f"Semi-finished recipes: {len(semi_finished)}")
for r in semi_finished:
    print(f"  {r['recipe_code']:8s} | {r['dish_name']:40s} | {r['ingredient_count']} ingredients")

print(f"\nFinished recipes: {len(finished)}")
for r in finished:
    sf_refs = [i['ingredient'] for i in r['ingredients'] if i['is_semi_finished_ref']]
    sf_info = f" (uses: {', '.join(sf_refs)})" if sf_refs else ""
    print(f"  {r['recipe_code']:8s} | {r['dish_name']:40s} | {r['ingredient_count']} ingredients{sf_info}")

# Unique ingredients across all recipes
all_ings = set()
for r in semi_finished + finished:
    for i in r['ingredients']:
        if not i['is_semi_finished_ref']:
            all_ings.add(i['ingredient'])
print(f"\nTotal unique raw ingredients: {len(all_ings)}")
