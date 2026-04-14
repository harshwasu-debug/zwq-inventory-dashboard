import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

with open(r"E:\Cloud Kitchen\AI Teams\canonical_price_list.json", "r") as f:
    prices_data = json.load(f)

price_map = {}
for item in prices_data["items"]:
    price_map[item["ingredient"].strip().lower()] = {
        "price": item["price_per_unit"], "uom": item["uom"]
    }

# User corrections
price_map["beef tenderloin (india)"] = {"price": 37.50, "uom": "Kg"}
price_map["chicken breast (aurora)"] = {"price": 13.00, "uom": "Kg"}
price_map["chicken thigh boneless"] = {"price": 13.00, "uom": "Kg"}
price_map["chicken wings uncooked"] = {"price": 13.00, "uom": "Kg"}
price_map["beef mince"] = {"price": 43.00, "uom": "Kg"}
price_map["tap water"] = {"price": 0, "uom": "L"}
price_map["korean soybean paste"] = {"price": 16.875, "uom": "Kg"}

alias = {
    "cooking oil": "oil vegetable",
    "boiled eggs": "egg",
    "arwa 500ml": "arwa water",
}

# N/A UOM = Piece for Indian recipes
def normalize_uom(uom):
    if uom in ("N/A", "n/a"):
        return "Piece"
    return uom

def gprice(name, uom="Kg", qty=1):
    n = name.strip().lower()
    if n in price_map:
        return price_map[n]["price"] * qty
    if n in alias and alias[n] in price_map:
        return price_map[alias[n]]["price"] * qty
    return None

wb = openpyxl.load_workbook(
    r"C:\Users\harsh\Desktop\Cloud Kitchen\Brands Home\Indian Brands\Recipes\Indian Recipes.xlsx",
    data_only=True
)

# SF recipes (row 3 has headers at row 3)
ws = wb["Semi Finished"]
sf = {}
for r in range(4, ws.max_row + 1):
    d = ws.cell(row=r, column=2).value
    i = ws.cell(row=r, column=3).value
    q = ws.cell(row=r, column=13).value
    u = ws.cell(row=r, column=14).value
    w = ws.cell(row=r, column=15).value
    if not d or not i:
        continue
    sf.setdefault(d, []).append({
        "i": str(i).strip(), "q": float(q or 0),
        "u": str(u or "Kg"), "w": float(w or 0)
    })

# Compute SF costs iteratively
sf_cost = {}
for _ in range(10):
    changed = False
    for name, ings in sf.items():
        if name in sf_cost:
            continue
        cost = 0; yld = 0; ok = True
        for ig in ings:
            q = ig["q"]; w = ig["w"] / 100
            g = q / (1 - w) if 0 < w < 1 else q
            sf_match = None
            for sn in sf_cost:
                if sn.lower() == ig["i"].lower():
                    sf_match = sn; break
            if sf_match:
                if ig["u"] in ("Kg", "L"):
                    cost += g * sf_cost[sf_match]["ppu"]
                else:
                    cost += g * sf_cost[sf_match]["total"]
            else:
                unresolved = any(sn.lower() == ig["i"].lower() and sn not in sf_cost for sn in sf)
                if unresolved:
                    ok = False; break
                p = gprice(ig["i"], ig["u"], g)
                if p is not None:
                    cost += p
            yld += q
        if ok:
            sf_cost[name] = {
                "total": round(cost, 4), "yield": round(yld, 4),
                "ppu": round(cost / yld, 4) if yld > 0 else cost,
            }
            changed = True
    if not changed:
        break

# Finished recipes from FnF Final
ws2 = wb["FnF Final"]
fin = {}; order = []; last_cat = ""
for r in range(3, ws2.max_row + 1):
    cat = ws2.cell(row=r, column=1).value
    d = ws2.cell(row=r, column=2).value
    i = ws2.cell(row=r, column=3).value
    q = ws2.cell(row=r, column=13).value
    u = ws2.cell(row=r, column=14).value
    w = ws2.cell(row=r, column=15).value
    if cat: last_cat = cat
    if not d or not i: continue
    if d not in fin:
        fin[d] = {"cat": last_cat, "ings": []}
        order.append(d)
    if not fin[d]["cat"]: fin[d]["cat"] = last_cat
    fin[d]["ings"].append({
        "i": str(i).strip(), "q": float(q or 0),
        "u": normalize_uom(str(u or "Kg")), "w": float(w or 0)
    })

# Read the old pre-computed costs for packaging/chutney/pickle values
ws3 = wb["final Costing"]
old_costs = {}
for r in range(2, ws3.max_row + 1):
    name = ws3.cell(row=r, column=2).value
    if not name: continue
    old_costs[str(name).strip().lower()] = {
        "mint": float(str(ws3.cell(row=r, column=4).value or 0).replace(",", "")),
        "pickle": float(str(ws3.cell(row=r, column=5).value or 0).replace(",", "")),
        "pkg": float(str(ws3.cell(row=r, column=6).value or 0).replace(",", "")),
    }

# Cost finished dishes - first pass
dish_cost_map = {}
results = []; missing_all = set()
for dn in order:
    data = fin[dn]; cost = 0; miss = []
    for ig in data["ings"]:
        q = ig["q"]; w = ig["w"] / 100
        g = q / (1 - w) if 0 < w < 1 else q
        il = ig["i"].lower()
        sf_match = None
        for sn in sf_cost:
            if sn.lower() == il: sf_match = sn; break
        if sf_match:
            if ig["u"] in ("Kg", "L"):
                cost += g * sf_cost[sf_match]["ppu"]
            else:
                cost += g * sf_cost[sf_match]["total"]
        else:
            p = gprice(ig["i"], ig["u"], g)
            if p is not None:
                cost += p
            else:
                miss.append(ig["i"]); missing_all.add(ig["i"])

    dish_cost_map[dn.lower().strip()] = cost

    old = old_costs.get(dn.lower().strip(), {"mint": 0, "pickle": 0, "pkg": 0})
    results.append({
        "dish": dn, "cat": data["cat"], "food_cost": round(cost, 2),
        "mint_chutney": old["mint"], "pickle": old["pickle"], "packaging": old["pkg"],
        "cost": round(cost + old["mint"] + old["pickle"] + old["pkg"], 2), "miss": miss
    })

# Second pass: resolve cross-dish references (Chicken Tikka in Butter Chicken etc.)
for idx, dn in enumerate(order):
    if not results[idx]["miss"]:
        continue
    data = fin[dn]; cost = 0; miss = []
    for ig in data["ings"]:
        q = ig["q"]; w = ig["w"] / 100
        g = q / (1 - w) if 0 < w < 1 else q
        il = ig["i"].lower()
        sf_match = None
        for sn in sf_cost:
            if sn.lower() == il: sf_match = sn; break
        if sf_match:
            if ig["u"] in ("Kg", "L"):
                cost += g * sf_cost[sf_match]["ppu"]
            else:
                cost += g * sf_cost[sf_match]["total"]
        elif il in dish_cost_map:
            # Cross-dish ref: cost per Kg of the referenced dish
            ref_cost = dish_cost_map[il]
            # Estimate: dish serves ~250g, so ppu = food_cost / 0.25
            ref_ppu = ref_cost / 0.25 if ref_cost > 0 else 0
            cost += g * ref_ppu
        else:
            p = gprice(ig["i"], ig["u"], g)
            if p is not None:
                cost += p
            else:
                miss.append(ig["i"])
    dish_cost_map[dn.lower().strip()] = cost
    old = old_costs.get(dn.lower().strip(), {"mint": 0, "pickle": 0, "pkg": 0})
    results[idx]["food_cost"] = round(cost, 2)
    results[idx]["cost"] = round(cost + old["mint"] + old["pickle"] + old["pkg"], 2)
    results[idx]["miss"] = miss

# Fix beverages
for r in results:
    if r["dish"] == "Mineral Water":
        r["food_cost"] = 0.44; r["cost"] = 0.44; r["miss"] = []

for i, r in enumerate(results, 1):
    m = " [!" + ",".join(set(r["miss"])) + "]" if r["miss"] else ""
    print(f'{i:>3d}. [{r["cat"]:<25s}] {r["dish"]:<45s} AED {r["cost"]:>8.2f} (food:{r["food_cost"]:.2f}){m}')

print(f"\nTotal: {len(results)} dishes")
if missing_all:
    print(f"Missing: {sorted(missing_all)}")

with open(r"E:\Cloud Kitchen\AI Teams\indian_dish_costing.json", "w") as f:
    json.dump({"dishes": results}, f, indent=2)
print("Saved to indian_dish_costing.json")
