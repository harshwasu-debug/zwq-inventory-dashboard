from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# Read all data from existing sheet
src = load_workbook(r'E:\Cloud Kitchen\AI Teams\Dish_Costing\Complete_Dish_Costing_Final.xlsx')
ws_src = src['American Menu']

data_rows = []
for r in range(2, ws_src.max_row + 1):
    sno = ws_src.cell(row=r, column=1).value
    cat = ws_src.cell(row=r, column=2).value
    dish = ws_src.cell(row=r, column=3).value
    cost = ws_src.cell(row=r, column=4).value
    sell = ws_src.cell(row=r, column=5).value
    if dish:
        data_rows.append((sno, cat, dish, cost, sell))

print(f"Read {len(data_rows)} dishes from source")

# Create fresh workbook
wb = Workbook()
ws = wb.active
ws.title = "American Menu"

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_font = Font(name='Arial', size=10)
data_font_blue = Font(name='Arial', size=10, color='0000FF')  # hardcoded inputs
data_font_black = Font(name='Arial', size=10, color='000000')  # formulas
currency_fmt = '#,##0.00'
pct_fmt = '0.0%'
int_fmt = '#,##0'

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

# Headers
headers = [
    ('S.No.', 6),
    ('Category', 28),
    ('Dish Name', 42),
    ('Cost (AED)', 12),
    ('Sell Price (AED)', 14),
    ('', 2),           # spacer col F
    ('Margin', 12),
    ('Margin%', 10),
    ('', 2),           # spacer col I
    ('Ideal Price', 12),
    ('', 2),           # spacer col K
    ('Final Price', 12),
]

for c, (name, width) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[cell.column_letter].width = width

# Freeze top row
ws.freeze_panes = 'A2'

# Data rows
for i, (sno, cat, dish, cost, sell) in enumerate(data_rows):
    r = i + 2

    # S.No.
    ws.cell(row=r, column=1, value=sno).font = data_font
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')

    # Category
    ws.cell(row=r, column=2, value=cat).font = data_font

    # Dish Name
    ws.cell(row=r, column=3, value=dish).font = data_font

    # Cost (blue = hardcoded input)
    c_cell = ws.cell(row=r, column=4, value=cost)
    c_cell.font = data_font_blue
    c_cell.number_format = currency_fmt

    # Sell Price (blue = hardcoded input)
    s_cell = ws.cell(row=r, column=5, value=sell)
    s_cell.font = data_font_blue
    s_cell.number_format = int_fmt

    # Margin formula (black = formula)
    m_cell = ws.cell(row=r, column=7)
    m_cell.value = f'=IFERROR(((E{r}-MIN(E{r}/2,30)-4)*0.7)-$D{r},0)'
    m_cell.font = data_font_black
    m_cell.number_format = currency_fmt

    # Margin%
    mp_cell = ws.cell(row=r, column=8)
    mp_cell.value = f'=IFERROR(G{r}/E{r},0)'
    mp_cell.font = data_font_black
    mp_cell.number_format = pct_fmt

    # Ideal Price
    ip_cell = ws.cell(row=r, column=10)
    ip_cell.value = f'=ROUND(IF(D{r}<=3.2,28+10*(D{r}),(23.8+D{r})/0.45),0)'
    ip_cell.font = data_font_black
    ip_cell.number_format = int_fmt

    # Final Price
    fp_cell = ws.cell(row=r, column=12)
    fp_cell.value = f'=MAX(J{r},E{r})'
    fp_cell.font = data_font_black
    fp_cell.number_format = int_fmt

    # Borders on all data cells
    for c in range(1, 13):
        ws.cell(row=r, column=c).border = thin_border

    # Alternate row shading
    if i % 2 == 1:
        light_fill = PatternFill('solid', fgColor='F2F2F2')
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = light_fill

# Auto-filter
ws.auto_filter.ref = f'A1:L{len(data_rows) + 1}'

# Save
output = r'E:\Cloud Kitchen\AI Teams\Dish_Costing\American_Menu_Costing.xlsx'
wb.save(output)
print(f"Saved fresh costing sheet to {output}")
print(f"Total dishes: {len(data_rows)}")
