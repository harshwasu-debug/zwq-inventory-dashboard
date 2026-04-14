import json
from openpyxl import load_workbook

with open(r'E:\Cloud Kitchen\AI Teams\Dish_Costing\american_dish_costing.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = load_workbook(r'E:\Cloud Kitchen\AI Teams\Dish_Costing\Complete_Dish_Costing_Final.xlsx', data_only=True)
ws = wb['American Menu']

sheet_data = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    name = row[2].value
    cost = row[3].value
    sell = row[4].value
    if name:
        sheet_data[name.strip().lower()] = {'name': name, 'cost': cost, 'sell': sell}

base_menu = [
    ('Classic Beef Burger', 'Bronx Beef Burgers', 66),
    ('Crunchy BBQ Burger', 'Bronx Beef Burgers', 72),
    ('Beef Slaw Burger', 'Bronx Beef Burgers', 68),
    ('Beef Burger Quesadilla', 'Bronx Beef Burgers', 67),
    ('Buffalo Beef Burger', 'Bronx Beef Burgers', 66),
    ('Texas Beef Burger', 'Bronx Beef Burgers', 66),
    ('Mushroom Truffle Beef Burger', 'Bronx Beef Burgers', 68),
    ('Beef Double Cheese Burger', 'Bronx Beef Burgers', 68),
    ('Double Beef Burger', 'Bronx Beef Burgers', 82),
    ('Bacon Beef Burger', 'Bronx Beef Burgers', 71),
    ('BBQ Crispy Beef Burger', 'Bronx Beef Burgers', 72),
    ('Mexican Crunch Beef Burger', 'Bronx Beef Burgers', 67),
    ('Grilled Classic Chicken Burger', 'Bronx Chicken Burgers', 53),
    ('Crispy Chicken Burger', 'Bronx Chicken Burgers', 55),
    ('Chicken Truffle Burger', 'Bronx Chicken Burgers', 58),
    ('Honey BBQ Chicken Burger', 'Bronx Chicken Burgers', 56),
    ('Chicken Crunchy Slaw Burger', 'Bronx Chicken Burgers', 56),
    ('Chicken Dynamite Burger', 'Bronx Chicken Burgers', 56),
    ('Chicken Honey Mustard Burger', 'Bronx Chicken Burgers', 56),
    ('Chicken Avocado Burger', 'Bronx Chicken Burgers', 57),
    ('Chicken Buffalo Burger', 'Bronx Chicken Burgers', 56),
    ('Mexican Crunch Chicken Burger', 'Bronx Chicken Burgers', 56),
    ('Hot Honey Burger', 'Bronx Chicken Burgers', 56),
    ('Asian Ranch Burger', 'Bronx Chicken Burgers', 56),
    ('Italian Chicken Burger', 'Bronx Chicken Burgers', 56),
    ('Chili Chicken Burger', 'Bronx Chicken Burgers', 56),
    ('Classic Beef Sliders', 'Bronx Beef Sliders', 64),
    ('Crunchy BBQ Sliders', 'Bronx Beef Sliders', 69),
    ('Beef Slaw Sliders', 'Bronx Beef Sliders', 66),
    ('Buffalo Beef Sliders', 'Bronx Beef Sliders', 64),
    ('Texas Beef Sliders', 'Bronx Beef Sliders', 64),
    ('Mushroom Truffle Beef Sliders', 'Bronx Beef Sliders', 66),
    ('Beef Double Cheese Sliders', 'Bronx Beef Sliders', 65),
    ('Double Beef Sliders', 'Bronx Beef Sliders', 78),
    ('Bacon Beef Sliders', 'Bronx Beef Sliders', 68),
    ('BBQ Crispy Beef Sliders', 'Bronx Beef Sliders', 69),
    ('Mexican Crunch Beef Sliders', 'Bronx Beef Sliders', 65),
    ('Grilled Classic Chicken Sliders', 'Bronx Chicken Sliders', 51),
    ('Crispy Chicken Sliders', 'Bronx Chicken Sliders', 55),
    ('Chicken Truffle Sliders', 'Bronx Chicken Sliders', 56),
    ('Honey BBQ Chicken Sliders', 'Bronx Chicken Sliders', 55),
    ('Chicken Crunchy Slaw Sliders', 'Bronx Chicken Sliders', 55),
    ('Chicken Dynamite Sliders', 'Bronx Chicken Sliders', 55),
    ('Chicken Honey Mustard Sliders', 'Bronx Chicken Sliders', 55),
    ('Chicken Avocado Sliders', 'Bronx Chicken Sliders', 57),
    ('Chicken Buffalo Sliders', 'Bronx Chicken Sliders', 56),
    ('Mexican Crunch Chicken Sliders', 'Bronx Chicken Sliders', 56),
    ('Hot Honey Sliders', 'Bronx Chicken Sliders', 55),
    ('Asian Ranch Sliders', 'Bronx Chicken Sliders', 55),
    ('Italian Chicken Sliders', 'Bronx Chicken Sliders', 54),
    ('Chili Chicken Sliders', 'Bronx Chicken Sliders', 56),
    ('Classic Grilled Halloumi Slider', 'Bronx Veggie Sliders', 52),
    ('Buffalo Grilled Halloumi Slider', 'Bronx Veggie Sliders', 53),
    ('BBQ Grilled Halloumi Slider', 'Bronx Veggie Sliders', 54),
    ('Honey Mustard Grilled Halloumi Slider', 'Bronx Veggie Sliders', 53),
    ('Avocado Grilled Halloumi Slider', 'Bronx Veggie Sliders', 55),
    ('Classic Falafel Slider', 'Bronx Veggie Sliders', 42),
    ('Buffalo Falafel Slider', 'Bronx Veggie Sliders', 44),
    ('BBQ Falafel Slider', 'Bronx Veggie Sliders', 45),
    ('Honey Mustard Falafel Slider', 'Bronx Veggie Sliders', 44),
    ('Avocado Falafel Slider', 'Bronx Veggie Sliders', 45),
    ('Chicken Tenders', 'Bronx Wings Tenders & Wraps', 59),
    ('Chicken Wings', 'Bronx Wings Tenders & Wraps', 57),
    ('Chicken Wraps', 'Bronx Wings Tenders & Wraps', 55),
    ('Chicken Burger Combo Meal for 1', 'Bronx Combo Meals', 66),
    ('Chicken Burger Combo Meal for 2', 'Bronx Combo Meals', 116),
    ('Chicken Burger Combo Meal for 4', 'Bronx Combo Meals', 178),
    ('Beef Burger Combo Meal for 1', 'Bronx Combo Meals', 72),
    ('Beef Burger Combo Meal for 2', 'Bronx Combo Meals', 124),
    ('Beef Burger Combo Meal for 4', 'Bronx Combo Meals', 198),
    ('Beef Sliders Meal for 1', 'Bronx Slider Meals', 72),
    ('Beef Sliders Meal for 2', 'Bronx Slider Meals', 128),
    ('Beef Sliders Meals for 4', 'Bronx Slider Meals', 194),
    ('Chicken Sliders Meal for 1', 'Bronx Slider Meals', 62),
    ('Chicken Sliders Meal for 2', 'Bronx Slider Meals', 108),
    ('Chicken Sliders Meal for 4', 'Bronx Slider Meals', 172),
    ('Fries', 'Sides', 19),
    ('Spicy Fries', 'Sides', 22),
    ('Dynamite Fries', 'Sides', 24),
    ('Cheesy Fries', 'Sides', 26),
    ('Tex Mex Fries', 'Sides', 30),
    ('Onion Rings', 'Sides', 24),
    ('Curly Fries', 'Sides', 22),
    ('Mozzarella Sticks', 'Sides', 29),
    ('Dynamite Shrimp', 'Sides', 44),
    ('Veggie Sticks', 'Sides', 19),
    ('Nachos', 'Sides', 34),
    ('Loaded Nachos', 'Sides', 52),
    ('House Slaw', 'Sides', 25),
    ('Garlic Aioli Dip', 'Dips & Sauces', 8),
    ('Sriracha Mayo Dip', 'Dips & Sauces', 8),
    ('Honey BBQ Dip', 'Dips & Sauces', 8),
    ('Dynamite Sauce Dip', 'Dips & Sauces', 8),
    ('Truffle Mayo Dip', 'Dips & Sauces', 8),
    ('Honey Mustard Dip', 'Dips & Sauces', 8),
    ('Mexican Ranch Dip', 'Dips & Sauces', 8),
    ('Cheese Dip', 'Dips & Sauces', 8),
    ('Avocado Ranch Dip', 'Dips & Sauces', 8),
    ('Mineral Water', 'Drinks', 4),
    ('Coca Cola', 'Drinks', 9),
    ('Coca Cola Zero', 'Drinks', 9),
    ('Sprite', 'Drinks', 9),
    ('Sprite Zero', 'Drinks', 9),
    ('Fanta', 'Drinks', 9),
]

aliases = {
    'classic beef burger': 'beef burger',
    'chicken wings': 'plain wings',
    'chicken tenders': 'plain tenders',
    'chicken wraps': 'classic chicken wrap',
    'coca cola': 'coca-cola',
    'coca cola zero': 'coca-cola zero',
    'chicken burger combo meal for 1': 'chicken burger combo (1 pax)',
    'chicken burger combo meal for 2': 'chicken burger combo (2 pax)',
    'chicken burger combo meal for 4': 'chicken burger combo (4 pax)',
    'beef burger combo meal for 1': 'beef burger combo (1 pax)',
    'beef burger combo meal for 2': 'beef burger combo (2 pax)',
    'beef burger combo meal for 4': 'beef burger combo (4 pax)',
    'beef sliders meals for 4': 'beef sliders meal for 4',
}

mismatches = []
missing_sell = []
missing_dish = []
ok = []

for name, cat, menu_price in base_menu:
    nl = name.lower()
    lookup = aliases.get(nl, nl)

    if lookup in sheet_data:
        sd = sheet_data[lookup]
        if sd['sell'] is None:
            missing_sell.append((name, cat, menu_price, sd['cost']))
        elif abs(sd['sell'] - menu_price) > 0.5:
            mismatches.append((name, cat, menu_price, sd['sell'], sd['cost']))
        else:
            ok.append((name, cat, menu_price, sd['cost']))
    else:
        missing_dish.append((name, cat, menu_price))

print(f"OK (matched & price correct): {len(ok)}")
print(f"Sell price mismatches: {len(mismatches)}")
print(f"Missing sell price: {len(missing_sell)}")
print(f"Not in costing sheet: {len(missing_dish)}")

if mismatches:
    print(f"\n=== SELL PRICE MISMATCHES ({len(mismatches)}) ===")
    for name, cat, mp, sp, cost in mismatches:
        print(f"  {name:45s} | Menu: {mp:>3} | Sheet: {int(sp):>3} | Cost: {cost:.2f}")

if missing_sell:
    print(f"\n=== MISSING SELL PRICE ({len(missing_sell)}) ===")
    for name, cat, mp, cost in missing_sell:
        print(f"  {name:45s} | Menu: {mp:>3} | Cost: {cost:.2f}")

if missing_dish:
    print(f"\n=== NOT IN COSTING SHEET ({len(missing_dish)}) ===")
    for name, cat, mp in missing_dish:
        print(f"  {name:45s} | Cat: {cat} | Menu Price: {mp}")

# Known costing bugs
print(f"\n=== KNOWN COSTING BUGS ===")
print(f"  Wings Meal for 1: shows 19.48, should be ~9.74 (duplicated from Meal for 2)")
print(f"  Chicken Sliders Meal for 1: shows 15.05, should be ~7.45 (duplicated from Meal for 2)")
