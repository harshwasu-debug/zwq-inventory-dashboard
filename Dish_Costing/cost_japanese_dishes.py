import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

with open(r"E:\Cloud Kitchen\AI Teams\canonical_price_list.json", "r") as f:
    prices_data = json.load(f)

price_map = {}
for item in prices_data["items"]:
    price_map[item["ingredient"].strip().lower()] = {
        "price": item["price_per_unit"], "uom": item["uom"]
    }

# User corrections (same as all other scripts)
price_map["beef tenderloin (india)"] = {"price": 37.50, "uom": "Kg"}
price_map["chicken breast (aurora)"] = {"price": 13.00, "uom": "Kg"}
price_map["chicken thigh boneless"] = {"price": 13.00, "uom": "Kg"}
price_map["chicken wings uncooked"] = {"price": 13.00, "uom": "Kg"}
price_map["beef mince"] = {"price": 43.00, "uom": "Kg"}
price_map["tap water"] = {"price": 0, "uom": "L"}
price_map["korean soybean paste"] = {"price": 16.875, "uom": "Kg"}

alias = {
    "cooking oil": "oil vegetable",
    "arwa 500ml": "arwa water",
}

# Piece-based items that need weight conversion
# Gyoza Wrapper: 140g pack = ~28 wrappers at 6.5/pack -> 0.232/wrapper
# Udon Noodles: 1 piece = 1 pack = 250g = 0.25Kg
# Egg Ramen Noodles: 1 piece = 1 pack = 200g = 0.20Kg
piece_overrides = {
    "gyoza wrapper": 0.232,        # AED per piece (wrapper)
    "udon noodles": 2.90,          # AED per piece (250g pack, 11.60/Kg)
    "egg ramen noodles": 2.95,     # AED per piece (200g pack, 14.75/Kg)
}

def gprice(name, uom="Kg", qty=1):
    n = name.strip().lower()
    # Piece override for items where "1 Piece" means a specific weight
    if uom == "Piece" and n in piece_overrides:
        return piece_overrides[n] * qty
    if n in price_map:
        return price_map[n]["price"] * qty
    if n in alias and alias[n] in price_map:
        return price_map[alias[n]]["price"] * qty
    return None

wb = openpyxl.load_workbook(
    r"C:\Users\harsh\Desktop\Cloud Kitchen\Brands Home\Japanese Food\Matt New Menu Recipes.xlsx",
    data_only=True
)

# SF recipes
ws = wb["Semi Finished"]
sf = {}
for r in range(3, ws.max_row + 1):
    d = ws.cell(row=r, column=2).value
    i = ws.cell(row=r, column=3).value
    q = ws.cell(row=r, column=13).value
    u = ws.cell(row=r, column=14).value
    w = ws.cell(row=r, column=15).value
    if not d or not i:
        continue
    sf.setdefault(d, []).append({
        "i": str(i).strip(), "q": float(q or 0),
        "u": str(u or "Kg"), "w": float(w or 0)
    })

# Compute SF costs iteratively
sf_cost = {}
for _ in range(10):
    changed = False
    for name, ings in sf.items():
        if name in sf_cost:
            continue
        cost = 0; yld = 0; ok = True
        for ig in ings:
            q = ig["q"]; w = ig["w"] / 100
            g = q / (1 - w) if 0 < w < 1 else q
            sf_match = None
            for sn in sf_cost:
                if sn.lower() == ig["i"].lower():
                    sf_match = sn; break
            if sf_match:
                if ig["u"] in ("Kg", "L"):
                    cost += g * sf_cost[sf_match]["ppu"]
                else:
                    cost += g * sf_cost[sf_match]["total"]
            else:
                unresolved = False
                for sn in sf:
                    if sn.lower() == ig["i"].lower() and sn not in sf_cost:
                        unresolved = True; break
                if unresolved:
                    ok = False; break
                p = gprice(ig["i"], ig["u"], g)
                if p is not None:
                    cost += p
            yld += q
        if ok:
            sf_cost[name] = {
                "total": round(cost, 4), "yield": round(yld, 4),
                "ppu": round(cost / yld, 4) if yld > 0 else cost,
            }
            changed = True
    if not changed:
        break

# Finished recipes
ws = wb["Finished Recipe"]
fin = {}; order = []; last_cat = ""
for r in range(3, ws.max_row + 1):
    cat = ws.cell(row=r, column=1).value
    d = ws.cell(row=r, column=2).value
    i = ws.cell(row=r, column=3).value
    q = ws.cell(row=r, column=13).value
    u = ws.cell(row=r, column=14).value
    w = ws.cell(row=r, column=15).value
    if cat: last_cat = cat
    if not d or not i: continue
    if d not in fin:
        fin[d] = {"cat": last_cat, "ings": []}
        order.append(d)
    if not fin[d]["cat"]: fin[d]["cat"] = last_cat
    fin[d]["ings"].append({
        "i": str(i).strip(), "q": float(q or 0),
        "u": str(u or "Kg"), "w": float(w or 0)
    })

# Cost finished dishes - first pass (individual dishes)
dish_cost_map = {}
results = []; missing_all = set()
for dn in order:
    data = fin[dn]; cost = 0; miss = []
    for ig in data["ings"]:
        q = ig["q"]; w = ig["w"] / 100
        g = q / (1 - w) if 0 < w < 1 else q
        il = ig["i"].lower()
        sf_match = None
        for sn in sf_cost:
            if sn.lower() == il: sf_match = sn; break
        if sf_match:
            if ig["u"] in ("Kg", "L"):
                cost += g * sf_cost[sf_match]["ppu"]
            else:
                cost += g * sf_cost[sf_match]["total"]
        else:
            p = gprice(ig["i"], ig["u"], g)
            if p is not None:
                cost += p
            else:
                miss.append(ig["i"]); missing_all.add(ig["i"])
    dish_cost_map[dn.lower().strip()] = cost
    results.append({"dish": dn, "cat": data["cat"], "cost": round(cost, 2), "miss": miss})

# Second pass: resolve combo/set meals that reference other finished dishes
for idx, dn in enumerate(order):
    data = fin[dn]
    if not results[idx]["miss"]:
        continue
    # Try resolving missing items from dish_cost_map
    cost = 0; miss = []
    for ig in data["ings"]:
        q = ig["q"]; w = ig["w"] / 100
        g = q / (1 - w) if 0 < w < 1 else q
        il = ig["i"].lower().strip()
        sf_match = None
        for sn in sf_cost:
            if sn.lower() == il: sf_match = sn; break
        if sf_match:
            if ig["u"] in ("Kg", "L"):
                cost += g * sf_cost[sf_match]["ppu"]
            else:
                cost += g * sf_cost[sf_match]["total"]
        elif il in dish_cost_map:
            # Reference to another finished dish (combo/set)
            cost += dish_cost_map[il] * g
        else:
            p = gprice(ig["i"], ig["u"], g)
            if p is not None:
                cost += p
            else:
                miss.append(ig["i"])
    results[idx]["cost"] = round(cost, 2)
    results[idx]["miss"] = miss
    dish_cost_map[dn.lower().strip()] = cost

missing_all = set()
for r in results:
    for m in r["miss"]:
        missing_all.add(m)

for i, r in enumerate(results, 1):
    m = " [!" + ",".join(set(r["miss"])) + "]" if r["miss"] else ""
    print(f'{i:>3d}. [{r["cat"]:<25s}] {r["dish"]:<45s} AED {r["cost"]:>8.2f}{m}')
print(f"\nTotal: {len(results)} dishes")
if missing_all:
    print(f"Missing: {sorted(missing_all)}")

with open(r"E:\Cloud Kitchen\AI Teams\japanese_dish_costing.json", "w") as f:
    json.dump({"dishes": results, "sf_costs": {k: v for k, v in sf_cost.items()}}, f, indent=2)
print("Saved to japanese_dish_costing.json")
