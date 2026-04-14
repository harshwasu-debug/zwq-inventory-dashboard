import json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load all data
with open(r"E:\Cloud Kitchen\AI Teams\korean_dish_costing.json", "r") as f:
    kd = json.load(f)
with open(r"E:\Cloud Kitchen\AI Teams\american_dish_costing.json", "r") as f:
    ad = json.load(f)
with open(r"E:\Cloud Kitchen\AI Teams\mexican_dish_costing.json", "r") as f:
    mxd = json.load(f)
with open(r"E:\Cloud Kitchen\AI Teams\indian_dish_costing.json", "r") as f:
    ind = json.load(f)
with open(r"E:\Cloud Kitchen\AI Teams\ramadan_combo_costing.json", "r") as f:
    kr_ram = json.load(f)
with open(r"E:\Cloud Kitchen\AI Teams\ramadan_combo_costing_all_brands.json", "r", encoding="utf-8") as f:
    am_ram = json.load(f)
with open(r"E:\Cloud Kitchen\AI Teams\mexican_ramadan_combo_costing.json", "r") as f:
    mx_ram = json.load(f)
with open(r"E:\Cloud Kitchen\AI Teams\indian_ramadan_combo_costing.json", "r") as f:
    in_ram = json.load(f)
with open(r"E:\Cloud Kitchen\AI Teams\all_menu_prices.json", "r") as f:
    menu_prices = json.load(f)

# Korean dish list from json (may be dict with 'dishes' key or list)
if isinstance(kd, dict) and "dishes" in kd:
    kd_list = kd["dishes"]
elif isinstance(kd, list):
    kd_list = kd
else:
    kd_list = []

# Styles
hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

def make_header(ws, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border = thin

def write_rows(ws, rows, start=2):
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            c = ws.cell(row=start + i, column=j + 1, value=val)
            c.border = thin
            if isinstance(val, float):
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
            elif isinstance(val, int) and j >= 2:
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")

def auto_width(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 55)

# Menu price lookup with aliases
def get_sell_price(dish_name, price_dict):
    nl = dish_name.strip().lower()
    aliases = {
        "cowboy beef burger": "crunchy bbq burger",
        "cowboy beef sliders": "crunchy bbq sliders",
        "chicken house slaw (portion) burger": "chicken house slaw burger",
        "chicken house slaw (portion) sliders": "chicken house slaw sliders",
        "honey mustard grilled halloumi": "honey mustard grilled halloumi slider",
        "mineral tap water": "mineral water",
        "coca cola": "coca-cola", "coca cola zero": "coca-cola zero",
        "coca-cola": "coca cola", "coca-cola zero": "coca cola zero",
        "coca cola can": "coca cola", "coca cola zero can": "coca cola zero",
        "fanta can": "fanta", "sprite can": "sprite", "sprite zero can": "sprite zero",
        "plain wings": "chicken wings", "plain tenders": "chicken tenders",
        "classic chicken wrap": "chicken wraps",
        "chicken burger combo (1 pax)": "chicken burger combo meal for 1",
        "chicken burger combo (2 pax)": "chicken burger combo meal for 2",
        "chicken burger combo (4 pax)": "chicken burger combo meal for 4",
        "beef burger combo (1 pax)": "beef burger combo meal for 1",
        "beef burger combo (2 pax)": "beef burger combo meal for 2",
        "beef burger combo (4 pax)": "beef burger combo meal for 4",
        "beef sliders meal for 4": "beef sliders meals for 4",
        "dynamite shrimps": "dynamite shrimp", "french fries": "fries",
        "cheese fries": "cheesy fries", "nachos and guacamole": "nachos",
        "cheesy nachos": "loaded nachos", "jarritos soda": "jarritos",
        "chicken burger": "chicken burger combo meal for 1",
        "mushroom beef burger": "mushroom truffle beef burger",
        "mushroom beef sliders": "mushroom truffle beef sliders",
    }
    p = price_dict.get(nl)
    if p is None and nl in aliases:
        p = price_dict.get(aliases[nl])
    return p

# Korean categories
k_cats = {
    "Beef Bibimbap": "Bibimbap", "Chicken Bibimbap": "Bibimbap", "Tofu Bibimbap": "Bibimbap",
    "Veg Bibimbap": "Bibimbap", "Shrimp Bibimbap": "Bibimbap",
    "Kimchi Shrimp Fried Rice": "Rice Bowls", "Bulgogi Deopbap": "Rice Bowls",
    "Dak Deopbap": "Rice Bowls", "Chili Garlic Shrimp Deopbap": "Rice Bowls",
    "Gochujang Shrimp Deopbap": "Rice Bowls", "Crispy Tofu Deopbap": "Rice Bowls",
    "Ojingeo Deopbap (Squid)": "Rice Bowls", "Nakji Bokkeum (Octopus)": "Rice Bowls",
    "Donkkaseu": "Rice Bowls",
    "Japchae Beef": "Japchae", "Japchae Chicken": "Japchae",
    "Japchae Veg": "Japchae", "Shrimp Japchae": "Japchae",
    "Jjajangmyeon": "Noodles",
    "Kimchi Jjigae": "Stews", "Soondubu Jjigae": "Stews",
    "Shrimp Soondubu Jjigae": "Stews", "Doenjang Jjigae": "Stews",
    "Yukgaejang": "Stews", "Jjamppong Bap": "Stews", "Jjamppong": "Stews",
    "Shin Ramen": "Ramen", "Beef Ramen": "Ramen", "Prawn Ramen": "Ramen",
    "Garlic Butter Shrimp Ramen": "Ramen", "Chicken Ramen": "Ramen",
    "Soy and Garlic Chicken": "Fried Chicken", "Yangnyeom Chicken": "Fried Chicken",
    "Korean Style Chicken": "Fried Chicken", "Spicy Chicken": "Fried Chicken",
    "Honey Garlic Chicken": "Fried Chicken",
    "Soy and Garlic Wings": "Wings", "Yangnyeom Wings": "Wings",
    "Spicy Wings": "Wings", "Honey Garlic Wings": "Wings",
    "Dynamite Shrimp": "Appetizers", "Rabokki": "Appetizers",
    "Tteokbokki with Cheese": "Appetizers", "Spicy Shrimp Tteokbokki": "Appetizers",
    "Deep Fried Chicken Mandu": "Appetizers", "Steamed Chicken Mandu": "Appetizers",
    "Gyeran-mari (Rolled Omelette)": "Appetizers",
    "Honey Chilli Shrimp": "Appetizers", "Honey Chilli Potatoes": "Appetizers",
    "Tuna Kimbap": "Kimbap", "Beef Kimbap": "Kimbap",
    "Cheese and Veg Kimbap": "Kimbap", "Chicken Kimbap": "Kimbap",
    "Kimchi Kimbap": "Kimbap", "Japchae & Shrimp Kimbap": "Kimbap",
    "House Kimchi": "Banchan", "Kongnamul Muchim": "Banchan",
    "Sigumchi Namul": "Banchan", "Danmuji (Pickled Radish)": "Banchan",
    "Rice Bowl Set": "Combo Meals", "Solo Chicken Set": "Combo Meals",
    "Chicken Combo for Two": "Combo Meals",
    "Mineral Tap Water": "Beverages", "Coca Cola": "Beverages",
    "Coca Cola Zero": "Beverages", "Fanta": "Beverages",
    "Sprite": "Beverages", "Sprite Zero": "Beverages",
}

wb = openpyxl.Workbook()

# ========== KOREAN MENU ==========
ws = wb.active
ws.title = "Korean Menu"
make_header(ws, ["S.No.", "Category", "Dish Name", "Cost (AED)", "Sell Price (AED)"])
rows = []
for i, d in enumerate(kd_list, 1):
    name = d.get("dish_name", d.get("dish", ""))
    cost = d.get("total_cost", d.get("cost", 0))
    cat = k_cats.get(name, "")
    sell = get_sell_price(name, menu_prices["korean"])
    rows.append([i, cat, name, float(cost), sell])
write_rows(ws, rows)
auto_width(ws)

# ========== KOREAN RAMADAN ==========
ws2 = wb.create_sheet("Korean Ramadan Combos")
make_header(ws2, ["S.No.", "Combo Name", "Sell Price (AED)", "Min Cost (AED)", "Max Cost (AED)"])
rows2 = [[i + 1, r["name"], int(r["price"]), r["min_cost"], r["max_cost"]] for i, r in enumerate(kr_ram)]
write_rows(ws2, rows2)
auto_width(ws2)

# ========== AMERICAN MENU ==========
ws3 = wb.create_sheet("American Menu")
make_header(ws3, ["S.No.", "Category", "Dish Name", "Cost (AED)", "Sell Price (AED)"])
rows3 = []
for i, d in enumerate(ad["dishes"], 1):
    sell = get_sell_price(d["dish"], menu_prices["american"])
    rows3.append([i, d["cat"], d["dish"], d["cost"], sell])
write_rows(ws3, rows3)
auto_width(ws3)

# ========== AMERICAN RAMADAN (3 sheets) ==========
burger_brands = ["Bronx Burger House", "Big Dawg's Burgers", "The Patty Pit", "Smashville Burgers"]
slider_brands = ["Juicy Buns", "The Slider Shack"]
wing_brands = ["Wingin' It", "Wings of Fury"]
combo_hdr = ["S.No.", "Brands", "Combo Name", "Sell Price (AED)", "Min Cost (AED)", "Max Cost (AED)"]

def dedup_combos(brand_list):
    seen = set()
    out = []
    sno = 0
    for r in am_ram:
        if r["brand"] not in brand_list:
            continue
        key = (r["combo"], r["sell"])
        if key in seen:
            continue
        seen.add(key)
        sno += 1
        out.append([sno, ", ".join(brand_list), r["combo"], int(r["sell"]), r["min_cost"], r["max_cost"]])
    return out

for sheet_name, brand_list in [
    ("Ramadan Burger Brands", burger_brands),
    ("Ramadan Slider Brands", slider_brands),
    ("Ramadan Wing Brands", wing_brands),
]:
    ws_r = wb.create_sheet(sheet_name)
    make_header(ws_r, combo_hdr)
    rows_r = dedup_combos(brand_list)
    write_rows(ws_r, rows_r)
    auto_width(ws_r)

# ========== MEXICAN MENU ==========
ws_mx = wb.create_sheet("Mexican Menu")
make_header(ws_mx, ["S.No.", "Category", "Dish Name", "Cost (AED)", "Sell Price (AED)"])
rows_mx = []
for i, d in enumerate(mxd["dishes"], 1):
    sell = get_sell_price(d["dish"], menu_prices["mexican"])
    rows_mx.append([i, d["cat"], d["dish"], d["cost"], sell])
write_rows(ws_mx, rows_mx)
auto_width(ws_mx)

# ========== MEXICAN RAMADAN ==========
ws_mxr = wb.create_sheet("Mexican Ramadan Combos")
mex_brands = "Loco Taco, Picante, MexiGo, Casa Del Queso, Fiesta"
make_header(ws_mxr, combo_hdr)
rows_mxr = [[i + 1, mex_brands, r["name"], int(r["sell"]), r["min_cost"], r["max_cost"]] for i, r in enumerate(mx_ram)]
write_rows(ws_mxr, rows_mxr)
auto_width(ws_mxr)

# ========== INDIAN MENU ==========
ws_in = wb.create_sheet("Indian Menu")
make_header(ws_in, ["S.No.", "Category", "Dish Name", "Food Cost", "Mint Chutney", "Pickle", "Packaging", "Total Cost (AED)", "Sell Price (AED)"])
rows_in = []
for i, d in enumerate(ind["dishes"], 1):
    sell = get_sell_price(d["dish"], menu_prices["indian"])
    rows_in.append([i, d["cat"], d["dish"], d["food_cost"], d["mint_chutney"], d["pickle"], d["packaging"], d["cost"], sell])
write_rows(ws_in, rows_in)
auto_width(ws_in)

# ========== INDIAN RAMADAN ==========
ws_inr = wb.create_sheet("Indian Ramadan Combos")
make_header(ws_inr, ["S.No.", "Brand", "Combo Name", "Sell Price (AED)", "Cost (AED)"])
rows_inr = [[i + 1, r["brand"], r["name"], int(r["sell"]), r["cost"]] for i, r in enumerate(in_ram)]
write_rows(ws_inr, rows_inr)
auto_width(ws_inr)

# Save
out_path = r"E:\Cloud Kitchen\AI Teams\Complete_Dish_Costing_Final.xlsx"
wb.save(out_path)

# Count totals
total = len(rows) + len(rows2) + len(rows3) + len(rows_mx) + len(rows_in)
total += sum(len(dedup_combos(bl)) for bl in [burger_brands, slider_brands, wing_brands])
total += len(rows_mxr) + len(rows_inr)

print(f"Saved: {out_path}")
print(f"  Korean Menu:            {len(rows)} dishes")
print(f"  Korean Ramadan:         {len(rows2)} combos")
print(f"  American Menu:          {len(rows3)} dishes")
print(f"  Ramadan Burger Brands:  {len(dedup_combos(burger_brands))} combos")
print(f"  Ramadan Slider Brands:  {len(dedup_combos(slider_brands))} combos")
print(f"  Ramadan Wing Brands:    {len(dedup_combos(wing_brands))} combos")
print(f"  Mexican Menu:           {len(rows_mx)} dishes")
print(f"  Mexican Ramadan:        {len(rows_mxr)} combos")
print(f"  Indian Menu:            {len(rows_in)} dishes")
print(f"  Indian Ramadan:         {len(rows_inr)} combos")
print(f"  TOTAL:                  {total} items across {len(wb.sheetnames)} sheets")

# Verify sell price coverage
for label, data_rows in [("Korean", rows), ("American", rows3), ("Mexican", rows_mx), ("Indian", rows_in)]:
    matched = sum(1 for r in data_rows if r[-1] is not None)
    print(f"  {label} sell prices: {matched}/{len(data_rows)}")
