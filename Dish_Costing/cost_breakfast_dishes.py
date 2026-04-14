import json, re, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RECIPES_PATH  = r'C:\Users\harsh\Desktop\Cloud Kitchen\Brands Home\Breakfast Brands\WIP files\recipes_for_ppt.json'
PRODUCTS_PATH = r'C:\Users\harsh\Desktop\Cloud Kitchen\Supy Upload\product_list.json'
OUTPUT_PATH   = r'C:\Users\harsh\Desktop\Cloud Kitchen\Brands Home\Breakfast Brands\WIP files\Before_Noon_Costing.xlsx'

with open(RECIPES_PATH,  'r', encoding='utf-8') as f:
    recipes = json.load(f)
with open(PRODUCTS_PATH, 'r', encoding='utf-8') as f:
    products = json.load(f)

# ── Parse buying unit string → quantity in base unit ──────────────────────────
def parse_buying_qty(buying_unit_str, base_unit):
    """e.g. '200ml' + 'L' → 0.2  |  '500g' + 'Kg' → 0.5  |  '1Ltr' + 'L' → 1.0"""
    s = str(buying_unit_str).strip()
    m = re.search(r'([\d,.]+)', s)
    if not m:
        return 1.0
    num = float(m.group(1).replace(',', ''))
    low = s.lower()
    if base_unit in ('L', 'Ltr', 'l'):
        if 'ml' in low:
            return num / 1000.0
        return num          # Ltr / L
    if base_unit in ('Kg', 'KG', 'kg'):
        if 'g' in low and 'kg' not in low:
            return num / 1000.0
        return num          # Kg
    # Piece / default
    return num

# ── Build price map: ingredient_lower → price_per_base_unit ──────────────────
# For multi-supplier ingredients, keep the cheapest price.
price_map = {}   # key: name_lower, val: {'price': float, 'unit': str, 'raw_name': str}

for p in products:
    name = str(p.get('Base Item / Ingredient Name') or '').strip()
    if not name:
        continue
    raw_price = p.get('Price Per Buying Unit')
    if raw_price is None:
        continue
    try:
        price_buy = float(raw_price)
    except (ValueError, TypeError):
        continue

    base_unit  = str(p.get('Base Unit') or 'Piece').strip()
    buying_str = str(p.get('Buying Unit') or '1').strip()
    qty_base   = parse_buying_qty(buying_str, base_unit)
    ppu        = price_buy / qty_base if qty_base else price_buy

    key = name.lower()
    if key not in price_map or ppu < price_map[key]['price']:
        price_map[key] = {'price': ppu, 'unit': base_unit, 'raw_name': name}

# Unit family compatibility for smarter matching
def unit_family(unit_str):
    u = (unit_str or '').upper().strip()
    if u in ('L', 'LTR', 'LITRE', 'LITER'):
        return 'volume'
    if u in ('KG', 'K', 'KILOGRAM'):
        return 'weight'
    if u in ('PIECE', 'PCS', 'PC', 'PIECES', 'NOS', 'NO', 'UNIT', 'UNITS'):
        return 'count'
    return 'other'

RECIPE_UOM_FAMILY = {
    'ML': 'volume', 'L': 'volume', 'LTR': 'volume',
    'GM': 'weight', 'KG': 'weight',
    'PIECE': 'count', 'SLICE': 'count', 'SLICES': 'count',
    'PCS': 'count', 'PC': 'count',
}

# ── Fuzzy match: recipe ingredient → price map key ────────────────────────────
def find_match(ingredient_name, recipe_uom=''):
    """Return (matched_key, price, base_unit) or (None, 0, '')"""
    needle = ingredient_name.lower().strip()
    needle_clean = re.sub(r'\s*\(.*?\)', '', needle).strip()
    req_family = RECIPE_UOM_FAMILY.get(recipe_uom.upper(), 'other')

    def compatible(key):
        return req_family == 'other' or unit_family(price_map[key]['unit']) == req_family

    # Exact match first
    if needle in price_map and compatible(needle):
        m = price_map[needle]
        return needle, m['price'], m['unit']
    if needle_clean in price_map and compatible(needle_clean):
        m = price_map[needle_clean]
        return needle_clean, m['price'], m['unit']

    # Partial match: unit-compatible preferred, then any
    best_key, best_len = None, 0
    fallback_key, fallback_len = None, 0
    for key in price_map:
        if needle_clean in key or key in needle_clean:
            if compatible(key):
                if len(key) > best_len:
                    best_key, best_len = key, len(key)
            else:
                if len(key) > fallback_len:
                    fallback_key, fallback_len = key, len(key)

    # Only use cross-unit fallback if the needle is very specific (>= 3 words after clean)
    # to avoid "pineapple" matching "jarritos pineapple" etc.
    if not best_key and fallback_key:
        words_in_needle = len(needle_clean.split())
        # Accept fallback only if needle is multi-word (specific enough) or exact key match
        if words_in_needle >= 2 or fallback_key == needle_clean:
            best_key = fallback_key

    chosen = best_key
    if chosen:
        m = price_map[chosen]
        return chosen, m['price'], m['unit']

    return None, 0.0, ''

# ── Unit conversion: recipe qty → base unit qty ───────────────────────────────
def convert_qty(qty_str, recipe_uom, base_unit):
    try:
        qty = float(qty_str)
    except (ValueError, TypeError):
        qty = 0.0
    r = recipe_uom.upper()
    b = base_unit.upper() if base_unit else ''
    # Weight
    if r == 'GM' and b in ('KG', 'K', 'KGS'):
        return qty / 1000.0
    if r == 'KG' and b in ('KG', 'K', 'KGS'):
        return qty
    # Volume
    if r == 'ML' and b in ('L', 'LTR', 'LITRE', 'LITER'):
        return qty / 1000.0
    if r in ('L', 'LTR') and b in ('L', 'LTR', 'LITRE', 'LITER'):
        return qty
    # Cross-unit: GM recipe vs L product (density ≈ 1 kg/L for condiments/sauces)
    if r == 'GM' and b in ('L', 'LTR', 'LITRE', 'LITER'):
        return qty / 1000.0
    if r == 'ML' and b in ('KG', 'K', 'KGS'):
        return qty / 1000.0
    # Piece / Slice / count
    if r in ('PIECE', 'PCS', 'PC', 'SLICE', 'SLICES'):
        return qty
    # Fallback: return qty as-is (may be unit mismatch like Piece→Kg)
    return qty

# ── Cost every ingredient in every dish ───────────────────────────────────────
summary_rows   = []   # (dish, category, food_cost, pkg_cost, total_cost, n_missing)
breakdown_rows = []   # (dish, type, ingredient, qty, uom, matched_name, price_pu, cost, status)
unmatched_set  = {}   # ingredient_name → set of dish names

for dish_name, dish_data in recipes.items():
    category = dish_data.get('category', '')
    food_cost = 0.0
    pkg_cost  = 0.0
    n_missing = 0

    for ing_type, key in [('Food', 'ingredients'), ('Packaging', 'packaging')]:
        for ing in dish_data.get(key, []):
            ing_name = ing.get('ingredient', '')
            qty_str  = ing.get('qty', '0')
            uom      = ing.get('uom', '')

            matched_key, price_pu, base_unit = find_match(ing_name, uom)
            conv_qty = convert_qty(qty_str, uom, base_unit)
            line_cost = conv_qty * price_pu if matched_key else 0.0
            status    = 'MATCHED' if matched_key else 'NOT FOUND'

            if not matched_key:
                n_missing += 1
                if ing_name not in unmatched_set:
                    unmatched_set[ing_name] = set()
                unmatched_set[ing_name].add(dish_name)

            display_name = price_map[matched_key]['raw_name'] if matched_key else ''

            breakdown_rows.append((
                dish_name, category, ing_type, ing_name,
                float(qty_str) if qty_str else 0, uom,
                display_name, round(price_pu, 4), round(conv_qty, 5), round(line_cost, 4),
                status
            ))

            if ing_type == 'Food':
                food_cost += line_cost
            else:
                pkg_cost  += line_cost

    total_cost = food_cost + pkg_cost
    summary_rows.append((dish_name, category, round(food_cost, 2), round(pkg_cost, 2), round(total_cost, 2), n_missing))

# Sort summary by total cost descending
summary_rows.sort(key=lambda x: x[4], reverse=True)

# ── Excel styles ──────────────────────────────────────────────────────────────
TERRACOTTA = 'CF6228'
CREAM      = 'F5EDD8'
LIGHT_RED  = 'FFD0D0'
SAGE_FILL  = 'D6E4DA'
WHITE      = 'FFFFFF'
DARK       = '2B2B2B'

hdr_font  = Font(bold=True, color=WHITE, size=11, name='Arial')
hdr_fill  = PatternFill(start_color=TERRACOTTA, end_color=TERRACOTTA, fill_type='solid')
sub_font  = Font(bold=True, color=DARK, size=10, name='Arial')
sub_fill  = PatternFill(start_color=CREAM, end_color=CREAM, fill_type='solid')
thin      = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)
num_fmt   = '#,##0.00'
pct_fmt   = '0.00%'

def hdr(ws, cols):
    for j, h in enumerate(cols, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font  = hdr_font
        c.fill  = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin
    ws.row_dimensions[1].height = 28

def autofit(ws, overrides=None):
    overrides = overrides or {}
    for col in ws.columns:
        letter = col[0].column_letter
        if letter in overrides:
            ws.column_dimensions[letter].width = overrides[letter]
            continue
        mx = max((len(str(c.value or '')) for c in col), default=6)
        ws.column_dimensions[letter].width = min(mx + 3, 48)

def freeze(ws, cell='A2'):
    ws.freeze_panes = cell

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Dish Costing Summary
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Dish Costing Summary'
hdr(ws1, ['#', 'Dish', 'Category', 'Food Cost (AED)', 'Pkg Cost (AED)', 'Total COGS (AED)', 'Unmatched Ingr.'])

for i, (dish, cat, fc, pc, tc, nm) in enumerate(summary_rows, 1):
    row = i + 1
    ws1.cell(row=row, column=1, value=i).border = thin
    ws1.cell(row=row, column=2, value=dish).border = thin
    ws1.cell(row=row, column=3, value=cat).border = thin

    for col, val in [(4, fc), (5, pc), (6, tc)]:
        c = ws1.cell(row=row, column=col, value=val)
        c.number_format = num_fmt
        c.alignment = Alignment(horizontal='right')
        c.border = thin

    c7 = ws1.cell(row=row, column=7, value=nm)
    c7.border = thin
    c7.alignment = Alignment(horizontal='center')
    if nm > 0:
        c7.fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type='solid')

# Totals row
tr = len(summary_rows) + 2
ws1.cell(row=tr, column=2, value='TOTAL / AVERAGE').font = Font(bold=True, name='Arial')
for col, vals in [
    (4, [r[2] for r in summary_rows]),
    (5, [r[3] for r in summary_rows]),
    (6, [r[4] for r in summary_rows]),
]:
    c = ws1.cell(row=tr, column=col, value=round(sum(vals)/len(vals), 2))
    c.number_format = num_fmt
    c.alignment = Alignment(horizontal='right')
    c.font = Font(bold=True, name='Arial')
    c.fill = PatternFill(start_color=CREAM, end_color=CREAM, fill_type='solid')
    c.border = thin
ws1.cell(row=tr, column=3, value='Average').font = Font(italic=True, name='Arial')

autofit(ws1, {'A': 5, 'B': 36, 'C': 20, 'D': 16, 'E': 16, 'F': 18, 'G': 15})
freeze(ws1)

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Ingredient Breakdown
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Ingredient Breakdown')
hdr(ws2, ['Dish', 'Category', 'Type', 'Recipe Ingredient', 'Qty', 'UOM',
          'Matched Product', 'Price/Unit (AED)', 'Converted Qty', 'Line Cost (AED)', 'Status'])

for i, row_data in enumerate(breakdown_rows, 2):
    (dish, cat, ing_type, ing_name, qty, uom,
     matched, price_pu, conv_qty, line_cost, status) = row_data

    vals = [dish, cat, ing_type, ing_name, qty, uom, matched, price_pu, conv_qty, line_cost, status]
    for j, v in enumerate(vals, 1):
        c = ws2.cell(row=i, column=j, value=v)
        c.border = thin
        if j in (5, 8, 9, 10):
            c.number_format = num_fmt
            c.alignment = Alignment(horizontal='right')

    if status == 'NOT FOUND':
        for j in range(1, 12):
            ws2.cell(row=i, column=j).fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type='solid')

    if ing_type == 'Packaging':
        for j in range(1, 12):
            existing = ws2.cell(row=i, column=j).fill
            if existing.fgColor.rgb in ('00000000', 'FFFFFFFF', WHITE):
                ws2.cell(row=i, column=j).fill = PatternFill(start_color=SAGE_FILL, end_color=SAGE_FILL, fill_type='solid')

autofit(ws2, {'A': 30, 'B': 18, 'C': 12, 'D': 36, 'E': 8, 'F': 8, 'G': 36, 'H': 16, 'I': 14, 'J': 14, 'K': 12})
freeze(ws2)

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Unmatched Ingredients
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Unmatched Ingredients')
hdr(ws3, ['Ingredient (from Recipe)', 'Used In (Dishes)', '# Dishes', 'Notes'])

unmatched_sorted = sorted(unmatched_set.items(), key=lambda x: -len(x[1]))
for i, (ing_name, dishes) in enumerate(unmatched_sorted, 2):
    ws3.cell(row=i, column=1, value=ing_name).border = thin
    ws3.cell(row=i, column=2, value=', '.join(sorted(dishes))).border = thin
    ws3.cell(row=i, column=3, value=len(dishes)).border = thin
    ws3.cell(row=i, column=4, value='').border = thin
    ws3.cell(row=i, column=1).fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type='solid')

autofit(ws3, {'A': 36, 'B': 80, 'C': 10, 'D': 40})
freeze(ws3)

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4 — Product Price Reference
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Product Price Reference')
hdr(ws4, ['Ingredient Name', 'Price / Base Unit (AED)', 'Base Unit'])

price_list = sorted(price_map.items(), key=lambda x: x[1]['raw_name'].lower())
for i, (key, val) in enumerate(price_list, 2):
    ws4.cell(row=i, column=1, value=val['raw_name']).border = thin
    c = ws4.cell(row=i, column=2, value=round(val['price'], 4))
    c.number_format = num_fmt
    c.alignment = Alignment(horizontal='right')
    c.border = thin
    ws4.cell(row=i, column=3, value=val['unit']).border = thin

autofit(ws4, {'A': 36, 'B': 22, 'C': 12})
freeze(ws4)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)

# ── Console summary ───────────────────────────────────────────────────────────
total_matched   = sum(1 for r in breakdown_rows if r[10] == 'MATCHED')
total_breakdown = len(breakdown_rows)
total_unmatched = len(unmatched_set)
dishes_with_gaps = sum(1 for r in summary_rows if r[5] > 0)

print(f"Saved: {OUTPUT_PATH}")
print(f"  Dishes costed   : {len(summary_rows)}")
print(f"  Ingredient lines: {total_breakdown}  ({total_matched} matched, {total_breakdown - total_matched} not found)")
print(f"  Unique unmatched: {total_unmatched} ingredients")
print(f"  Dishes with gaps: {dishes_with_gaps}")
print()
print(f"{'Dish':<40s}  {'Food':>7s}  {'Pkg':>7s}  {'Total':>7s}  {'Miss':>4s}")
print("-" * 70)
for dish, cat, fc, pc, tc, nm in summary_rows:
    warn = f" *** {nm} missing" if nm else ""
    print(f"  {dish:<38s}  {fc:>7.2f}  {pc:>7.2f}  {tc:>7.2f}  {nm:>4d}{warn}")

if unmatched_set:
    print()
    print("UNMATCHED INGREDIENTS:")
    for ing, dishes in sorted(unmatched_set.items(), key=lambda x: -len(x[1])):
        print(f"  {ing:<40s}  ({len(dishes)} dishes)")
