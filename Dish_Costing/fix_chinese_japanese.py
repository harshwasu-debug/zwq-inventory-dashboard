import json, sys, io, openpyxl, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# PART 1: Fix Chinese menu-to-recipe matching
# ============================================================
with open(r"E:\Cloud Kitchen\AI Teams\chinese_dish_costing.json", "r") as f:
    cd = json.load(f)

recipe_costs = {}
for r in cd["finished_costs"]:
    recipe_costs[r["dish"].lower().strip()] = r["cost"]

# Build comprehensive menu->recipe name mapping
# Strategy: for each menu item, try to find the matching recipe
def find_chinese_recipe_cost(menu_dish, variant):
    ml = menu_dish.lower().strip()

    # Direct match
    if ml in recipe_costs:
        return recipe_costs[ml]

    # Strip variant suffixes
    clean = re.sub(r'\s*\((veg|chicken|shrimp)\)\s*$', '', ml, flags=re.IGNORECASE).strip()
    clean = re.sub(r'\s*(veg|chicken|shrimp)\s*$', '', clean, flags=re.IGNORECASE).strip()
    # Remove extra spaces
    clean = re.sub(r'\s+', ' ', clean).strip()

    if clean in recipe_costs:
        return recipe_costs[clean]

    # Variant-specific recipe name patterns
    if variant == "Veg":
        # Try: "Veg X", "X", menu name directly
        for pattern in [ml, clean, f"veg {clean}", clean.replace("non veg ", "veg ").replace("non ", "")]:
            if pattern in recipe_costs:
                return recipe_costs[pattern]
        # Paneer variant for veg starters
        paneer = clean.replace("veg ", "paneer ").replace("honey chilli veg", "honey chilli potato")
        if paneer in recipe_costs:
            return recipe_costs[paneer]
        # Starters: "Veg 65" -> "Paneer 65"
        if "65" in clean:
            if "paneer 65" in recipe_costs:
                return recipe_costs["paneer 65"]

    elif variant == "Chicken":
        # Try: "Chicken X", "Chk X", "X Chicken"
        for pattern in [
            f"chicken {clean}", f"chk {clean}",
            clean.replace("non ", "chicken ").replace("veg ", "chicken "),
            f"kung pao chicken" if "kung pao" in clean else "",
            f"honey chilli chicken" if "honey chilli" in clean else "",
            f"teriyaki chicken" if "teriyaki" in clean else "",
            f"black pepper chicken" if "black pepper" in clean else "",
            f"dynamite chicken" if "dynamite" in clean else "",
            f"tempura chicken" if "tempura" in clean else "",
            f"stir fried chk chilli" if "stir fried" in clean and "chilli" in clean else "",
            f"mountain chilli chk" if "mountain" in clean else "",
            f"hot garlic chicken" if "hot garlic" in clean else "",
            f"chicken 65" if "65" in clean else "",
            f"chk manchurian dry" if "manchurian" in clean and "gravy" not in clean else "",
            f"chk manchurian gravy" if "manchurian" in clean and "gravy" in clean else "",
            f"chicken kung pao sauce" if "kung pao" in clean and "gravy" in clean else "",
            f"chilli chicken gravy" if "chilli" in clean and "gravy" in clean else "",
            f"chk black pepper sauce" if "black pepper" in clean and "gravy" in clean else "",
            f"chicken schezwan sauce" if "schezwan" in clean and "gravy" in clean else "",
            f"chk oyster chilli sauce" if "oyster" in clean and "gravy" in clean else "",
            f"mongolian chicken" if "mongolian" in clean else "",
        ]:
            if pattern and pattern in recipe_costs:
                return recipe_costs[pattern]

    elif variant == "Shrimp":
        for pattern in [
            f"shrimp {clean}", f"prawns {clean}",
            f"dynamite shrimp" if "dynamite" in clean else "",
            f"tempura shrimp" if "tempura" in clean else "",
            f"chilli garlic prawns" if "chilli garlic" in clean or "hot garlic" in clean else "",
            f"black pepper prawns" if "black pepper" in clean else "",
            f"teriyaki prawns" if "teriyaki" in clean else "",
            f"mountain chilli prawn" if "mountain" in clean else "",
            f"kung pao prawns" if "kung pao" in clean else "",
            f"honey chilli shrimp" if "honey chilli" in clean else "",
        ]:
            if pattern and pattern in recipe_costs:
                return recipe_costs[pattern]

    # Mains gravy variants
    if "gravy" in ml or "(gravy)" in ml:
        base = clean.replace("(gravy)", "").replace("gravy", "").strip()
        base = re.sub(r'^non\s+', '', base).strip()
        for pattern in [
            f"veg {base} gravy" if variant == "Veg" else "",
            f"paneer chilli gravy" if "chilli" in base and variant == "Veg" else "",
            f"veg schezwan gravy" if "schezwan" in base and variant == "Veg" else "",
            f"veg manchurian gravy" if "manchurian" in base and variant == "Veg" else "",
            f"veg kung pao sauce" if "kung pao" in base and variant == "Veg" else "",
        ]:
            if pattern and pattern in recipe_costs:
                return recipe_costs[pattern]

    # Momos - map to dumplings
    if "momos" in ml:
        base = clean.replace("momos", "").strip()
        for pattern in [
            f"{base} chk dumplings" if variant == "Chicken" else "",
            f"{base} veg dumplings" if variant == "Veg" else "",
            f"spicy 65 chk momos" if "65" in base and variant == "Chicken" else "",
            f"spicy 65 veg momos" if "65" in base and variant == "Veg" else "",
            f"schezwan chk momos" if "schezwan" in base and variant == "Chicken" else "",
            f"schezwan veg momos" if "schezwan" in base and variant == "Veg" else "",
        ]:
            if pattern and pattern in recipe_costs:
                return recipe_costs[pattern]

    # Rice & Noodles - base recipe is veg, chicken/shrimp just adds protein cost
    if variant in ("Chicken", "Shrimp"):
        if clean in recipe_costs:
            protein_add = 1.30 if variant == "Chicken" else 3.00  # ~100g chicken/shrimp
            return recipe_costs[clean] + protein_add

    # Spring rolls
    if "spring rolls" in ml:
        if "chicken" in ml or variant == "Chicken":
            return recipe_costs.get("chk spring rolls (8pc)", 0)
        return recipe_costs.get("veg spring rolls (8pc)", 0)

    # Beverages
    bev_map = {
        "water 500 ml": recipe_costs.get("steamed rice", 0.44),  # placeholder
        "coca cola 300 ml": 2.04, "coca cola zero 300 ml": 2.04,
        "sprite 300 ml": 2.04, "sprite zero 300 ml": 2.04, "fanta 300 ml": 2.04,
    }
    if ml in bev_map:
        return bev_map[ml]

    return 0

# Apply matching
for d in cd["dishes"]:
    cost = find_chinese_recipe_cost(d["dish"], d["variant"])
    d["cost"] = round(cost, 2)

# Count matches
matched = sum(1 for d in cd["dishes"] if d["cost"] > 0)
print(f"Chinese menu: {matched}/{len(cd['dishes'])} items now have costs")

# Save
with open(r"E:\Cloud Kitchen\AI Teams\chinese_dish_costing.json", "w") as f:
    json.dump(cd, f, indent=2)

# ============================================================
# PART 2: Fix Japanese sell prices and Ramadan combos
# ============================================================
with open(r"E:\Cloud Kitchen\AI Teams\japanese_dish_costing.json", "r") as f:
    jd = json.load(f)
with open(r"E:\Cloud Kitchen\AI Teams\all_menu_prices.json", "r") as f:
    mp = json.load(f)
with open(r"E:\Cloud Kitchen\AI Teams\japanese_ramadan_snare.json", "r", encoding="utf-8") as f:
    snare = json.load(f)

# Japanese dish cost map
jdc = {}
for d in jd["dishes"]:
    jdc[d["dish"].lower().strip()] = d["cost"]

# Fix Oneesan Ramadan - all brands use same recipes
brand_map = {
    "Oneesan": {
        "Oneesan's Solace (Solo)": "the solo play",
        "Shared Harmony (Date)": "date night box",
        "Simple Comforts (Fast)": "the power feed",
        "The Family Gathering (Party)": "the crowd pleaser",
        "Sister's Favorites (Premium)": "signature flex",
        "The Artisan's Palette (Raw)": "raw bar feast",
        "The Warm Welcome (Crowd)": "the party starter",
    },
    "Hikari": {
        "The Solo Play": "the solo play",
        "Date Night Box": "date night box",
        "The Power Feed": "the power feed",
        "The Crowd Pleaser": "the crowd pleaser",
        "Signature Flex": "signature flex",
        "Raw Bar Feast": "raw bar feast",
        "The Party Starter": "the party starter",
    },
    "Norii": {
        "The Solo Edit": "the solo play",
        "The Rendezvous": "date night box",
        "The Daily Ritual": "the power feed",
        "The Socialite": "the crowd pleaser",
        "Norii Gold Collection": "signature flex",
        "The Sashimi Standard": "raw bar feast",
        "The After Hours": "the party starter",
    },
}

jr_results = []
for brand, combos in snare.items():
    for combo_name, data in combos.items():
        sell = data["price"]
        # Normalize non-breaking spaces
        clean_name = combo_name.replace("\xa0", " ").strip()
        recipe_key = brand_map.get(brand, {}).get(clean_name, clean_name.lower())
        cost = jdc.get(recipe_key, 0)
        jr_results.append({
            "brand": brand, "name": combo_name, "sell": sell, "cost": round(cost, 2),
        })

with open(r"E:\Cloud Kitchen\AI Teams\japanese_ramadan_combo_costing.json", "w") as f:
    json.dump(jr_results, f, indent=2)

print("\nJapanese Ramadan combos:")
for r in jr_results:
    margin = r["sell"] - r["cost"]
    pct = (margin / r["sell"] * 100) if r["sell"] > 0 else 0
    print(f"  [{r['brand']:<10s}] {r['name']:<45s} Sell:{r['sell']:>4.0f} Cost:{r['cost']:>6.2f} Margin:{pct:.0f}%")

# ============================================================
# PART 3: Rebuild workbook with fixes
# ============================================================
wb = openpyxl.load_workbook(r"E:\Cloud Kitchen\AI Teams\Complete_Dish_Costing_Final.xlsx")

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

def make_header(ws, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center"); c.border = thin

def auto_width(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 55)

# Delete old Chinese sheets
for sn in ["Chinese Recipes", "Chinese Menu"]:
    if sn in wb.sheetnames:
        del wb[sn]

# Chinese Recipes sheet (89 costed recipes)
ws1 = wb.create_sheet("Chinese Recipes")
make_header(ws1, ["S.No.", "Category", "Dish Name", "Cost (AED)"])
for i, r in enumerate(cd["finished_costs"], 1):
    for j, val in enumerate([i, r["cat"], r["dish"], r["cost"]], 1):
        c = ws1.cell(row=i+1, column=j, value=val)
        c.border = thin
        if isinstance(val, float):
            c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
auto_width(ws1)

# Chinese Menu sheet (158 items with costs + sell prices)
ws2 = wb.create_sheet("Chinese Menu")
make_header(ws2, ["S.No.", "Category", "Dish Name", "Variant", "Cost (AED)", "Sell Price (AED)"])
for i, d in enumerate(cd["dishes"], 1):
    cost_val = d["cost"] if d["cost"] > 0 else None
    for j, val in enumerate([i, d["cat"], d["dish"], d["variant"], cost_val, d["sell"]], 1):
        c = ws2.cell(row=i+1, column=j, value=val)
        c.border = thin
        if isinstance(val, float):
            c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
        elif isinstance(val, int) and j >= 5:
            c.number_format = "#,##0"; c.alignment = Alignment(horizontal="right")
auto_width(ws2)

# Fix Japanese Ramadan sheet
if "Japanese Ramadan Combos" in wb.sheetnames:
    del wb["Japanese Ramadan Combos"]
ws3 = wb.create_sheet("Japanese Ramadan Combos")
make_header(ws3, ["S.No.", "Brand", "Combo Name", "Sell Price (AED)", "Cost (AED)"])
for i, r in enumerate(jr_results, 1):
    for j, val in enumerate([i, r["brand"], r["name"], int(r["sell"]), r["cost"]], 1):
        c = ws3.cell(row=i+1, column=j, value=val)
        c.border = thin
        if isinstance(val, float):
            c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
auto_width(ws3)

# Also verify Japanese Menu sell prices
ws_jm = wb["Japanese Menu"]
jap_prices = mp.get("japanese", {})
# Check if sell price column exists
if ws_jm.cell(row=1, column=5).value != "Sell Price (AED)":
    c = ws_jm.cell(row=1, column=5, value="Sell Price (AED)")
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center"); c.border = thin

jap_matched = 0
for r in range(2, ws_jm.max_row + 1):
    name = ws_jm.cell(row=r, column=3).value
    if not name:
        continue
    nl = str(name).strip().lower()
    p = jap_prices.get(nl)
    if p:
        ws_jm.cell(row=r, column=5, value=p).border = thin
        ws_jm.cell(row=r, column=5).number_format = "#,##0"
        jap_matched += 1
auto_width(ws_jm)
print(f"\nJapanese Menu sell prices matched: {jap_matched}")

wb.save(r"E:\Cloud Kitchen\AI Teams\Complete_Dish_Costing_Final.xlsx")
print(f"\nSaved workbook. Sheets: {wb.sheetnames}")
