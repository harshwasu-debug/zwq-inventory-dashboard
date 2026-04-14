import json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open(r'E:\Cloud Kitchen\AI Teams\korean_dish_costing.json', 'r') as f:
    kd = json.load(f)['dishes']
with open(r'E:\Cloud Kitchen\AI Teams\american_dish_costing.json', 'r') as f:
    ad = json.load(f)['dishes']
with open(r'E:\Cloud Kitchen\AI Teams\ramadan_combo_costing.json', 'r') as f:
    kr = json.load(f)
with open(r'E:\Cloud Kitchen\AI Teams\ramadan_combo_costing_all_brands.json', 'r', encoding='utf-8') as f:
    ar = json.load(f)

# Korean categories
k_cats = {
    'Beef Bibimbap':'Bibimbap','Chicken Bibimbap':'Bibimbap','Tofu Bibimbap':'Bibimbap',
    'Veg Bibimbap':'Bibimbap','Shrimp Bibimbap':'Bibimbap',
    'Kimchi Shrimp Fried Rice':'Rice Bowls','Bulgogi Deopbap':'Rice Bowls',
    'Dak Deopbap':'Rice Bowls','Chili Garlic Shrimp Deopbap':'Rice Bowls',
    'Gochujang Shrimp Deopbap':'Rice Bowls','Crispy Tofu Deopbap':'Rice Bowls',
    'Ojingeo Deopbap (Squid)':'Rice Bowls','Nakji Bokkeum (Octopus)':'Rice Bowls','Donkkaseu':'Rice Bowls',
    'Japchae Beef':'Japchae','Japchae Chicken':'Japchae','Japchae Veg':'Japchae','Shrimp Japchae':'Japchae',
    'Jjajangmyeon':'Noodles',
    'Kimchi Jjigae':'Stews','Soondubu Jjigae':'Stews','Shrimp Soondubu Jjigae':'Stews',
    'Doenjang Jjigae':'Stews','Yukgaejang':'Stews','Jjamppong Bap':'Stews','Jjamppong':'Stews',
    'Shin Ramen':'Ramen','Beef Ramen':'Ramen','Prawn Ramen':'Ramen',
    'Garlic Butter Shrimp Ramen':'Ramen','Chicken Ramen':'Ramen',
    'Soy and Garlic Chicken':'Fried Chicken','Yangnyeom Chicken':'Fried Chicken',
    'Korean Style Chicken':'Fried Chicken','Spicy Chicken':'Fried Chicken','Honey Garlic Chicken':'Fried Chicken',
    'Soy and Garlic Wings':'Wings','Yangnyeom Wings':'Wings','Spicy Wings':'Wings','Honey Garlic Wings':'Wings',
    'Dynamite Shrimp':'Appetizers','Rabokki':'Appetizers','Tteokbokki with Cheese':'Appetizers',
    'Spicy Shrimp Tteokbokki':'Appetizers','Deep Fried Chicken Mandu':'Appetizers',
    'Steamed Chicken Mandu':'Appetizers','Gyeran-mari (Rolled Omelette)':'Appetizers',
    'Honey Chilli Shrimp':'Appetizers','Honey Chilli Potatoes':'Appetizers',
    'Tuna Kimbap':'Kimbap','Beef Kimbap':'Kimbap','Cheese and Veg Kimbap':'Kimbap',
    'Chicken Kimbap':'Kimbap','Kimchi Kimbap':'Kimbap','Japchae & Shrimp Kimbap':'Kimbap',
    'House Kimchi':'Banchan','Kongnamul Muchim':'Banchan','Sigumchi Namul':'Banchan',
    'Danmuji (Pickled Radish)':'Banchan',
    'Rice Bowl Set':'Combo Meals','Solo Chicken Set':'Combo Meals','Chicken Combo for Two':'Combo Meals',
    'Mineral Tap Water':'Beverages','Coca Cola':'Beverages','Coca Cola Zero':'Beverages',
    'Fanta':'Beverages','Sprite':'Beverages','Sprite Zero':'Beverages',
}

# Styles
hdr_font = Font(bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
num_fmt = '#,##0.00'
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def make_header(ws, headers):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
        c.border = thin

def write_data(ws, rows, start=2):
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            c = ws.cell(row=start+i, column=j+1, value=val)
            c.border = thin
            if isinstance(val, float):
                c.number_format = num_fmt
                c.alignment = Alignment(horizontal='right')

def auto_width(ws):
    for col in ws.columns:
        mx = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(mx + 3, 50)

wb = openpyxl.Workbook()

# === Sheet 1: Korean Menu ===
ws = wb.active
ws.title = 'Korean Menu'
make_header(ws, ['S.No.', 'Category', 'Dish Name', 'Cost (AED)'])
rows = [[i+1, k_cats.get(d['dish_name'], ''), d['dish_name'], d['total_cost']] for i, d in enumerate(kd)]
write_data(ws, rows)
auto_width(ws)

# === Sheet 2: Korean Ramadan Combos ===
ws2 = wb.create_sheet('Korean Ramadan Combos')
make_header(ws2, ['S.No.', 'Combo Name', 'Sell Price (AED)', 'Min Cost (AED)', 'Max Cost (AED)'])
rows2 = [[i+1, r['name'], float(r['price']), r['min_cost'], r['max_cost']] for i, r in enumerate(kr)]
write_data(ws2, rows2)
auto_width(ws2)

# === Sheet 3: American Menu ===
ws3 = wb.create_sheet('American Menu')
make_header(ws3, ['S.No.', 'Category', 'Dish Name', 'Cost (AED)'])
rows3 = [[i+1, d['cat'], d['dish'], d['cost']] for i, d in enumerate(ad)]
write_data(ws3, rows3)
auto_width(ws3)

# === Sheet 4: Ramadan Burger Brands ===
burger_brands = ['Bronx Burger House', "Big Dawg's Burgers", 'The Patty Pit', 'Smashville Burgers']
slider_brands = ['Juicy Buns', 'The Slider Shack']
wing_brands = ["Wingin' It", 'Wings of Fury']

combo_headers = ['S.No.', 'Brands', 'Combo Name', 'Sell Price (AED)', 'Min Cost (AED)', 'Max Cost (AED)']

def dedup_combos(brand_list):
    seen = set()
    out = []
    sno = 0
    for r in ar:
        if r['brand'] not in brand_list:
            continue
        key = (r['combo'], r['sell'])
        if key in seen:
            continue
        seen.add(key)
        sno += 1
        out.append([sno, ', '.join(brand_list), r['combo'], float(r['sell']), r['min_cost'], r['max_cost']])
    return out

ws4 = wb.create_sheet('Ramadan Burger Brands')
make_header(ws4, combo_headers)
rows4 = dedup_combos(burger_brands)
write_data(ws4, rows4)
auto_width(ws4)

ws5 = wb.create_sheet('Ramadan Slider Brands')
make_header(ws5, combo_headers)
rows5 = dedup_combos(slider_brands)
write_data(ws5, rows5)
auto_width(ws5)

ws6 = wb.create_sheet('Ramadan Wing Brands')
make_header(ws6, combo_headers)
rows6 = dedup_combos(wing_brands)
write_data(ws6, rows6)
auto_width(ws6)

# Save
out = r'E:\Cloud Kitchen\AI Teams\Complete_Dish_Costing.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'  Korean Menu: {len(rows)} dishes')
print(f'  Korean Ramadan Combos: {len(rows2)} combos')
print(f'  American Menu: {len(rows3)} dishes')
print(f'  Ramadan Burger Brands: {len(rows4)} combos')
print(f'  Ramadan Slider Brands: {len(rows5)} combos')
print(f'  Ramadan Wing Brands: {len(rows6)} combos')
total = len(rows) + len(rows2) + len(rows3) + len(rows4) + len(rows5) + len(rows6)
print(f'  TOTAL: {total} items across 6 sheets')
