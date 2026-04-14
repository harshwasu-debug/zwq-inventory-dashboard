from openpyxl import load_workbook
import csv

# ============================================================
# RENAME MAPPINGS
# ============================================================
renames = {
    'Plain Wings': 'Chicken Wings',
    'Plain Tenders': 'Chicken Tenders',
    'Classic Chicken Wrap': 'Chicken Wraps',
    'Beef Burger Combo (1 Pax)': 'Beef Burger Combo Meal for 1',
    'Beef Burger Combo (2 Pax)': 'Beef Burger Combo Meal for 2',
    'Beef Burger Combo (4 Pax)': 'Beef Burger Combo Meal for 4',
    'Chicken Burger Combo (1 Pax)': 'Chicken Burger Combo Meal for 1',
    'Chicken Burger Combo (2 Pax)': 'Chicken Burger Combo Meal for 2',
    'Chicken Burger Combo (4 Pax)': 'Chicken Burger Combo Meal for 4',
    'Coca-Cola': 'Coca Cola',
    'Coca-Cola Zero': 'Coca Cola Zero',
    'Honey Mustard Grilled Halloumi': 'Honey Mustard Grilled Halloumi Slider',
}

# ============================================================
# STEP 1: Update Recipe Sheet (American Recipes.xlsx)
# ============================================================
recipe_path = r'C:/Users/harsh/Desktop/Cloud Kitchen/Brands Home/American Brands/American Recipes.xlsx'
wb_recipe = load_workbook(recipe_path)
ws_fin = wb_recipe['Finished']

renamed_recipe = 0
for r in range(3, ws_fin.max_row + 1):
    cell = ws_fin.cell(row=r, column=2)
    if cell.value and cell.value.strip() in renames:
        old = cell.value.strip()
        cell.value = renames[old]
        renamed_recipe += 1

print(f"[Recipe Sheet - Finished] Renamed {renamed_recipe} cell entries")

# Count unique dishes after rename
dishes_after = set()
for r in range(3, ws_fin.max_row + 1):
    v = ws_fin.cell(row=r, column=2).value
    if v:
        dishes_after.add(v.strip())
print(f"[Recipe Sheet - Finished] {len(dishes_after)} unique dishes")

# ============================================================
# STEP 2: Add 3 additional recipes to Recipe Sheet
# ============================================================
additional_recipes = [
    ('Sides & Dips', 'Veggie Sticks', 'Carrot', 0.050, 'Kg', 15),
    ('Sides & Dips', 'Veggie Sticks', 'Cucumber', 0.060, 'Kg', 10),
    ('Sides & Dips', 'Veggie Sticks', 'Red Bell Pepper', 0.035, 'Kg', 12),
    ('Sides & Dips', 'Veggie Sticks', 'Yellow Bell Pepper', 0.035, 'Kg', 12),
    ('Sides & Dips', 'Veggie Sticks', 'Orange Bell Pepper', 0.025, 'Kg', 12),
    ('Sides & Dips', 'Veggie Sticks', 'Ranch Dressing', 0.045, 'Kg', 0),
    ('Sides & Dips', 'Veggie Sticks', 'Salad Bowl Packaging (Set)', 1, 'N/A', 0),
    ('Sides & Dips', 'Veggie Sticks', 'Final Bagging (Set)', 1, 'N/A', 0),
    ('Sides & Dips', 'Honey Chilli Potatoes', 'Potato', 0.15, 'Kg', 15),
    ('Sides & Dips', 'Honey Chilli Potatoes', 'Cornflour', 0.015, 'Kg', 0),
    ('Sides & Dips', 'Honey Chilli Potatoes', 'Cooking Oil', 0.03, 'L', 0),
    ('Sides & Dips', 'Honey Chilli Potatoes', 'Honey', 0.03, 'Kg', 2),
    ('Sides & Dips', 'Honey Chilli Potatoes', 'Chilli Flakes', 0.005, 'Kg', 0),
    ('Sides & Dips', 'Honey Chilli Potatoes', 'Soy Sauce', 0.005, 'L', 0),
    ('Sides & Dips', 'Honey Chilli Potatoes', 'Spring Onion', 0.005, 'Kg', 12),
    ('Sides & Dips', 'Honey Chilli Potatoes', 'Salad Bowl Packaging (Set)', 1, 'N/A', 0),
    ('Sides & Dips', 'Honey Chilli Potatoes', 'Final Bagging (Set)', 1, 'N/A', 0),
    ('Sides & Dips', 'Honey Chilli Shrimp', 'Shrimp', 0.15, 'Kg', 5),
    ('Sides & Dips', 'Honey Chilli Shrimp', 'Honey', 0.03, 'Kg', 2),
    ('Sides & Dips', 'Honey Chilli Shrimp', 'Chilli Flakes', 0.005, 'Kg', 0),
    ('Sides & Dips', 'Honey Chilli Shrimp', 'Soy Sauce', 0.005, 'L', 0),
    ('Sides & Dips', 'Honey Chilli Shrimp', 'Cooking Oil', 0.03, 'L', 0),
    ('Sides & Dips', 'Honey Chilli Shrimp', 'Spring Onion', 0.005, 'Kg', 12),
    ('Sides & Dips', 'Honey Chilli Shrimp', 'Salad Bowl Packaging (Set)', 1, 'N/A', 0),
    ('Sides & Dips', 'Honey Chilli Shrimp', 'Final Bagging (Set)', 1, 'N/A', 0),
]

# Check if these dishes already exist
existing = set()
for r in range(3, ws_fin.max_row + 1):
    v = ws_fin.cell(row=r, column=2).value
    if v:
        existing.add(v.strip())

new_dishes = set(row[1] for row in additional_recipes)
to_add = [d for d in new_dishes if d not in existing]

if to_add:
    # Find the structure: col A=Category, B=Dish, C=Ingredient, D=Qty, E=UOM, F=Wastage
    # Need to figure out which columns map to what
    # From exploration: col1=Category, col2=Dish, col3=Ingredient, col4=Qty, col5=UOM, col6=Wastage
    # But costing script uses: col13=Qty, col14=UOM, col15=Wastage
    # Let me check the actual column layout
    print(f"\n[Recipe Sheet] Headers row 2:")
    for c in range(1, 16):
        v = ws_fin.cell(row=2, column=c).value
        if v:
            print(f"  Col {c}: {v}")

    # Add recipes at the end
    last_row = ws_fin.max_row
    added_count = 0
    for cat, dish, ing, qty, uom, wastage in additional_recipes:
        if dish not in to_add:
            continue
        last_row += 1
        ws_fin.cell(row=last_row, column=1).value = cat if added_count == 0 or additional_recipes[[r[1] for r in additional_recipes].index(dish)][1] != additional_recipes[[r[1] for r in additional_recipes].index(dish) - 1][1] else None
        # Set category only on first ingredient of each dish
        ws_fin.cell(row=last_row, column=1).value = cat
        ws_fin.cell(row=last_row, column=2).value = dish
        ws_fin.cell(row=last_row, column=3).value = ing
        ws_fin.cell(row=last_row, column=4).value = qty
        ws_fin.cell(row=last_row, column=5).value = uom
        ws_fin.cell(row=last_row, column=6).value = wastage
        # Also set cols 13-15 (used by costing script)
        ws_fin.cell(row=last_row, column=13).value = qty
        ws_fin.cell(row=last_row, column=14).value = uom
        ws_fin.cell(row=last_row, column=15).value = wastage
        added_count += 1

    print(f"[Recipe Sheet] Added {added_count} ingredient rows for {len(to_add)} new dishes: {to_add}")
else:
    print("[Recipe Sheet] All 3 additional dishes already exist - skipping")

wb_recipe.save(recipe_path)
print(f"[Recipe Sheet] Saved to {recipe_path}")

# ============================================================
# STEP 3: Update Costing Sheet
# ============================================================
costing_path = r'E:\Cloud Kitchen\AI Teams\Dish_Costing\Complete_Dish_Costing_Final.xlsx'
wb_cost = load_workbook(costing_path)
ws_cost = wb_cost['American Menu']

# Apply renames to column C (Dish Name)
# Also rename "Beef Burger" -> "Classic Beef Burger" if it's a standalone entry
costing_renames = dict(renames)
# Check if "Beef Burger" exists separately from "Classic Beef Burger"
has_classic = False
has_plain = False
for r in range(2, ws_cost.max_row + 1):
    v = ws_cost.cell(row=r, column=3).value
    if v:
        if v.strip() == 'Classic Beef Burger':
            has_classic = True
        if v.strip() == 'Beef Burger':
            has_plain = True

if has_plain and has_classic:
    print("[Costing] Both 'Beef Burger' and 'Classic Beef Burger' exist - not renaming to avoid duplicate")
elif has_plain and not has_classic:
    costing_renames['Beef Burger'] = 'Classic Beef Burger'

renamed_cost = 0
for r in range(2, ws_cost.max_row + 1):
    cell = ws_cost.cell(row=r, column=3)
    if cell.value and cell.value.strip() in costing_renames:
        old = cell.value.strip()
        cell.value = costing_renames[old]
        renamed_cost += 1

print(f"[Costing Sheet] Renamed {renamed_cost} entries")

# Verify all 104 base menu items are present
base_menu_items = {
    'Classic Beef Burger': 66, 'Crunchy BBQ Burger': 72, 'Beef Slaw Burger': 68,
    'Beef Burger Quesadilla': 67, 'Buffalo Beef Burger': 66, 'Texas Beef Burger': 66,
    'Mushroom Truffle Beef Burger': 68, 'Beef Double Cheese Burger': 68,
    'Double Beef Burger': 82, 'Bacon Beef Burger': 71, 'BBQ Crispy Beef Burger': 72,
    'Mexican Crunch Beef Burger': 67, 'Grilled Classic Chicken Burger': 53,
    'Crispy Chicken Burger': 55, 'Chicken Truffle Burger': 58,
    'Honey BBQ Chicken Burger': 56, 'Chicken Crunchy Slaw Burger': 56,
    'Chicken Dynamite Burger': 56, 'Chicken Honey Mustard Burger': 56,
    'Chicken Avocado Burger': 57, 'Chicken Buffalo Burger': 56,
    'Mexican Crunch Chicken Burger': 56, 'Hot Honey Burger': 56,
    'Asian Ranch Burger': 56, 'Italian Chicken Burger': 56, 'Chili Chicken Burger': 56,
    'Classic Beef Sliders': 64, 'Crunchy BBQ Sliders': 69, 'Beef Slaw Sliders': 66,
    'Buffalo Beef Sliders': 64, 'Texas Beef Sliders': 64,
    'Mushroom Truffle Beef Sliders': 66, 'Beef Double Cheese Sliders': 65,
    'Double Beef Sliders': 78, 'Bacon Beef Sliders': 68, 'BBQ Crispy Beef Sliders': 69,
    'Mexican Crunch Beef Sliders': 65, 'Grilled Classic Chicken Sliders': 51,
    'Crispy Chicken Sliders': 55, 'Chicken Truffle Sliders': 56,
    'Honey BBQ Chicken Sliders': 55, 'Chicken Crunchy Slaw Sliders': 55,
    'Chicken Dynamite Sliders': 55, 'Chicken Honey Mustard Sliders': 55,
    'Chicken Avocado Sliders': 57, 'Chicken Buffalo Sliders': 56,
    'Mexican Crunch Chicken Sliders': 56, 'Hot Honey Sliders': 55,
    'Asian Ranch Sliders': 55, 'Italian Chicken Sliders': 54, 'Chili Chicken Sliders': 56,
    'Classic Grilled Halloumi Slider': 52, 'Buffalo Grilled Halloumi Slider': 53,
    'BBQ Grilled Halloumi Slider': 54, 'Honey Mustard Grilled Halloumi Slider': 53,
    'Avocado Grilled Halloumi Slider': 55, 'Classic Falafel Slider': 42,
    'Buffalo Falafel Slider': 44, 'BBQ Falafel Slider': 45,
    'Honey Mustard Falafel Slider': 44, 'Avocado Falafel Slider': 45,
    'Chicken Tenders': 59, 'Chicken Wings': 57, 'Chicken Wraps': 55,
    'Chicken Burger Combo Meal for 1': 66, 'Chicken Burger Combo Meal for 2': 116,
    'Chicken Burger Combo Meal for 4': 178, 'Beef Burger Combo Meal for 1': 72,
    'Beef Burger Combo Meal for 2': 124, 'Beef Burger Combo Meal for 4': 198,
    'Beef Sliders Meal for 1': 72, 'Beef Sliders Meal for 2': 128,
    'Beef Sliders Meal for 4': 194, 'Chicken Sliders Meal for 1': 62,
    'Chicken Sliders Meal for 2': 108, 'Chicken Sliders Meal for 4': 172,
    'Fries': 19, 'Spicy Fries': 22, 'Dynamite Fries': 24, 'Cheesy Fries': 26,
    'Tex Mex Fries': 30, 'Onion Rings': 24, 'Curly Fries': 22,
    'Mozzarella Sticks': 29, 'Dynamite Shrimp': 44, 'Veggie Sticks': 19,
    'Nachos': 34, 'Loaded Nachos': 52, 'House Slaw': 25,
    'Garlic Aioli Dip': 8, 'Sriracha Mayo Dip': 8, 'Honey BBQ Dip': 8,
    'Dynamite Sauce Dip': 8, 'Truffle Mayo Dip': 8, 'Honey Mustard Dip': 8,
    'Mexican Ranch Dip': 8, 'Cheese Dip': 8, 'Avocado Ranch Dip': 8,
    'Mineral Water': 4, 'Coca Cola': 9, 'Coca Cola Zero': 9,
    'Sprite': 9, 'Sprite Zero': 9, 'Fanta': 9,
}

# Build current sheet index
sheet_index = {}
for r in range(2, ws_cost.max_row + 1):
    name = ws_cost.cell(row=r, column=3).value
    if name:
        sheet_index[name.strip()] = r

# Check and fix
missing = []
price_fixed = 0
for menu_name, sell_price in base_menu_items.items():
    if menu_name in sheet_index:
        r = sheet_index[menu_name]
        current_sell = ws_cost.cell(row=r, column=5).value
        if current_sell is None or abs(current_sell - sell_price) > 0.5:
            ws_cost.cell(row=r, column=5).value = sell_price
            price_fixed += 1
    else:
        missing.append(menu_name)

print(f"[Costing Sheet] Fixed {price_fixed} remaining sell price mismatches")

if missing:
    print(f"[Costing Sheet] Still missing {len(missing)} items: {missing}")

wb_cost.save(costing_path)
print(f"[Costing Sheet] Saved to {costing_path}")

# Final summary
print(f"\n{'='*60}")
print(f"SUMMARY")
print(f"{'='*60}")
print(f"Recipe Sheet: {renamed_recipe} entries renamed, {len(to_add) if to_add else 0} new dishes added")
print(f"Costing Sheet: {renamed_cost} entries renamed, {price_fixed} sell prices fixed")
if missing:
    print(f"WARNING: {len(missing)} menu items still missing from costing sheet")
else:
    print(f"All 104 base menu items present in costing sheet")
