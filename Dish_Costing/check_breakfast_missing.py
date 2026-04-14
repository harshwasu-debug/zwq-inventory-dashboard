import openpyxl, json, sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Load canonical price list
with open(r'E:\Cloud Kitchen\AI Teams\canonical_price_list.json', 'r') as f:
    prices = json.load(f)

price_names = {}
for item in prices['items']:
    price_names[item['ingredient'].strip().lower()] = item['ingredient']

# Load breakfast recipes (Sheet2)
wb = openpyxl.load_workbook(
    r'C:\Users\harsh\Desktop\Cloud Kitchen\Brands Home\Breakfast Brands\Breakfast Brands Recipe.xlsx',
    data_only=True
)
ws = wb['Sheet2']

raw_ingredients = {}
for r in range(3, ws.max_row + 1):
    ing = ws.cell(row=r, column=4).value
    dish = ws.cell(row=r, column=3).value
    uom = ws.cell(row=r, column=6).value
    qty = ws.cell(row=r, column=5).value
    if not ing or not dish:
        continue
    raw = str(ing).strip()
    if raw not in raw_ingredients:
        raw_ingredients[raw] = {'dishes': set(), 'uom': str(uom or ''), 'qty_example': qty}
    raw_ingredients[raw]['dishes'].add(str(dish).strip())

# Normalize: strip parenthetical context
def normalize(name):
    n = name.strip()
    n = re.sub(r'\s*\(.*?\)\s*', ' ', n).strip()
    n = re.sub(r'\s+', ' ', n)
    return n.lower().strip()

# Build comprehensive mapping of breakfast ingredient -> canonical name
manual_map = {
    # Butter variants
    'butter (for cooking)': 'butter unsalted',
    'butter (for eggs)': 'butter unsalted',
    'butter (for mushrooms)': 'butter unsalted',
    'butter (for pan)': 'butter unsalted',
    'butter (for toast)': 'butter unsalted',
    'butter (for toasting)': 'butter unsalted',
    # Olive oil variants
    'olive oil': 'olive oil',
    'olive oil (cold-pressed drizzle)': 'olive oil',
    'olive oil (drizzle)': 'olive oil',
    'olive oil (for cooking)': 'olive oil',
    'olive oil (for grilling)': 'olive oil',
    'olive oil (for salad)': 'olive oil',
    'olive oil (for tapenade)': 'olive oil',
    'olive oil / lemon dressing': 'olive oil',
    # Egg variants
    'egg (scrambled)': 'egg',
    'egg (fried)': 'egg',
    'egg (poached)': 'egg',
    'egg (poached in sauce)': 'egg',
    'egg (for batter)': 'egg',
    'egg white (separated from 4 eggs)': 'egg',
    # Sriracha mayo
    'sriracha mayo': 'sriracha mayo',
    'sriracha mayo (drizzle)': 'sriracha mayo',
    'sriracha mayo (drizzle garnish)': 'sriracha mayo',
    'sriracha mayo (side)': 'sriracha mayo',
    # Chilli
    'chilli flakes (garnish)': 'chilli flakes',
    'chilli flakes / jalapeño': 'chilli flakes',
    # Jalapeno
    'jalapeno pickled (garnish)': 'jalapeno sliced',
    'jalapeno sliced (garnish)': 'jalapeno sliced',
    'jalapeno sliced (amer. classic)': 'jalapeno sliced',
    # Banana
    'banana fresh': 'banana',
    'banana fresh (coin slices)': 'banana',
    'banana fresh (sliced)': 'banana',
    # Herbs
    'chives (garnish)': 'chives',
    'chives (optional garnish)': 'chives',
    'flat-leaf parsley (garnish)': 'parsley',
    'fresh parsley (garnish)': 'parsley',
    'fresh cilantro': 'coriander leaves',
    'fresh cilantro (garnish)': 'coriander leaves',
    'fresh coriander (garnish)': 'coriander leaves',
    'fresh basil leaves (garnish)': 'basil leaves',
    'fresh mint leaves (garnish)': 'mint leaves',
    'fresh thyme / oregano': 'thyme',
    'dill leaves': 'dill leaves',
    'dill leaves (garnish)': 'dill leaves',
    # Spices/seasonings
    'black pepper (cracked)': 'black pepper powder',
    'black pepper (optional)': 'black pepper powder',
    'smoked paprika (garnish dusting)': 'paprika powder',
    'smoked paprika / aleppo pepper (added — missing from original)': 'paprika powder',
    'cinnamon (for batter)': 'cinnamon powder',
    'flaky sea salt': 'salt',
    'dukkah (garnish)': 'dukkah',
    # Vegetables
    'cherry tomato (finely diced)': 'cherry tomato',
    'cherry tomato (halved)': 'cherry tomato',
    'cherry tomato (halved, roasted)': 'cherry tomato',
    'tomato (grilled)': 'tomato',
    'mushroom white (sautéed)': 'mushroom white',
    'spinach leaves': 'spinach',
    'cucumber (julienned)': 'cucumber',
    'cucumber (sliced)': 'cucumber',
    'bell pepper mixed': 'capsicum red',
    'red onion (finely diced)': 'onion red',
    'red onion (for pickling)': 'onion red',
    'onion (diced)': 'onion white',
    'potato (diced & pre-cooked)': 'potato',
    'broccoli (steamed)': 'broccoli',
    'spring onion sliced (garnish)': 'spring onion',
    'spring onion slivers (garnish)': 'spring onion',
    'rocca lettuce': 'rocket leaves',
    'mixed greens': 'lettuce iceberg',
    'side salad greens (lollo rosso/iceberg)': 'lettuce iceberg',
    'avocado hass': 'avocado',
    # Fruits
    'lemon wedge': 'lemon',
    'lemon zest (garnish)': 'lemon',
    'lime wedge': 'lime',
    'fresh strawberry': 'strawberry',
    'strawberry (fresh/frozen, halved)': 'strawberry',
    'strawberry (frozen/fresh)': 'strawberry',
    'blueberry (fresh/frozen)': 'blueberry',
    'blueberry (frozen/fresh)': 'blueberry',
    'blueberry / grapes': 'blueberry',
    'raspberry (fresh/frozen)': 'raspberry',
    'kiwi (cubed)': 'kiwi',
    'kiwi (sliced into rounds)': 'kiwi',
    'mango (cubed)': 'mango',
    'mango (fresh/frozen, sliced)': 'mango',
    'pineapple (fresh/frozen, chunks)': 'pineapple',
    'watermelon (cubed)': 'watermelon',
    'green apple (cubed)': 'apple green',
    'red apple (cubed)': 'apple',
    'pomegranate seeds': 'pomegranate',
    'pomegranate seeds (added — missing from original)': 'pomegranate',
    'medjool dates (chopped)': 'dates medjool',
    'passion fruit (drizzle garnish)': 'passion fruit',
    # Cheese
    'cheese cheddar sliced': 'cheese slices yellow',
    'cheese slice': 'cheese slices yellow',
    'cheddar cheese shredded': 'cheese cheddar shredded',
    'cheddar shredded (garnish)': 'cheese cheddar shredded',
    'cream cheese (philadelphia)': 'cream cheese',
    'feta cheese (3 cow)': 'feta cheese',
    'halloumi cheese (green hill)': 'halloumi cheese',
    'fresh mozzarella (pinar/maestrella)': 'mozzarella cheese',
    # Dairy
    'milk (full fat)': 'milk full cream',
    'milk full fat (for batter)': 'milk full cream',
    'yoghurt (zad catering)': 'yoghurt',
    'sour cream (garnish)': 'sour cream',
    'sour cream (side)': 'sour cream',
    'ghee (amul/aseel)': 'ghee',
    # Sauces/condiments
    'lemon juice (anti-browning for apple)': 'lemon juice',
    'lemon juice (anti-browning for banana)': 'lemon juice',
    'lime juice': 'lime juice',
    'sriracha chili sauce': 'sriracha sauce',
    'ketchup (side)': 'ketchup',
    'tomato paste (anna) / peeled tomato': 'tomato paste',
    'balsamic vinegar (mara balsamic)': 'balsamic vinegar',
    'white vinegar (pickling liquid)': 'vinegar white',
    'honey (drizzle)': 'honey',
    'honey (optional drizzle)': 'honey',
    'vanilla extract (for batter)': 'vanilla extract',
    'maple syrup (taste basics honey)': 'maple syrup',
    'grain mustard (side)': 'mustard',
    'powdered sugar (garnish dusting)': 'sugar icing',
    'powdered sugar (optional garnish)': 'sugar icing',
    'sesame seeds (garnish)': 'sesame seeds white',
    'sesame seeds (on bun, garnish)': 'sesame seeds white',
    'olive black sliced (blended with olive oil as tapenade)': 'olives black sliced',
    'olives black': 'olives black sliced',
    # Bread/carbs
    'brioche bun': 'brioche bun 4.5 inch sliced',
    'brioche bun (sliced)': 'brioche bun 4.5 inch sliced',
    'brioche bun (sliced thick)': 'brioche bun 4.5 inch sliced',
    'brioche bun (sliced, half portion)': 'brioche bun 4.5 inch sliced',
    'brioche bun (cut into soldiers)': 'brioche bun 4.5 inch sliced',
    'sourdough bread': 'sourdough bread',
    'sourdough bread (side)': 'sourdough bread',
    'sourdough toast': 'sourdough bread',
    '10" flour tortilla': 'tortilla 10 inch',
    'corn tortilla 6"': 'corn tortilla 6 inch',
    'all-purpose flour': 'all purpose flour (maida)',
    'rice basmati (cooked)': 'basmati rice',
    'rice / quinoa base': 'basmati rice',
    # Protein
    'beef mince 80/20': 'beef mince',
    'tuna (kingway tuna in water)': 'tuna canned',
    'smoked salmon (frozen)': 'smoked salmon',
    'crispy beef bacon': 'beef bacon slices',
    'crispy beef bacon (frozen slices)': 'beef bacon slices',
    'americana spicy chicken strip': 'chicken tenderloin breaded',
    # Water
    'water': 'tap water',
    # Packaging
    'sauce cup 2oz clear with lid': 'sauce cup 4oz clear with lid',
    # Misc
    'truffle oil (black truffle oil)': 'truffle oil',
    'chipotle powder / chipotle in adobo': 'chipotle peppers',
    'peanut butter (drizzle)': 'peanut butter',
    'almond butter (drizzle)': 'almond butter',
    'toasted coconut flakes': 'coconut desiccated',
    'toasted coconut flakes (garnish)': 'coconut desiccated',
    'lotus biscoff biscuit (crushed garnish)': 'lotus biscoff biscuit',
    'lotus biscoff biscuit (crushed, optional garnish)': 'lotus biscoff biscuit',
    'pistachios (crushed)': 'pistachio',
    'pumpkin seeds (garnish)': 'pumpkin seeds',
    'hemp seeds (garnish)': 'hemp seeds',
    'cacao nibs': 'cacao nibs',
    'chia seeds': 'chia seeds',
    'chia seeds (garnish)': 'chia seeds',
    'microgreens (garnish)': 'microgreens',
    'jam / honey (optional side)': 'honey',
    'guacamole (side)': 'guacamole',
    'pico de gallo (tomato/onion/cilantro/jalapeno)': 'pico de gallo',
    'cheese sauce (cheddar-based) (added — missing from original)': 'cheddar cheese sauce',
    'gochujang sauce (added — missing from original)': 'gochujang',
    'harissa paste (added — missing from original)': 'harissa paste',
    'caramelised onion': 'caramelised onion',
}

# Check each ingredient
matched = {}
unmatched = {}

for raw_name, data in raw_ingredients.items():
    raw_lower = raw_name.lower().strip()

    # 1. Direct match
    if raw_lower in price_names:
        matched[raw_name] = price_names[raw_lower]
        continue

    # 2. Manual map
    if raw_lower in manual_map:
        target = manual_map[raw_lower].lower()
        if target in price_names:
            matched[raw_name] = price_names[target]
            continue

    # 3. Normalized (strip parentheses)
    norm = normalize(raw_name).lower()
    if norm in price_names:
        matched[raw_name] = price_names[norm]
        continue

    # 4. Check manual map with normalized
    if norm in manual_map:
        target = manual_map[norm].lower()
        if target in price_names:
            matched[raw_name] = price_names[target]
            continue

    # Not found
    unmatched[raw_name] = data

# Print results
print(f"Total unique ingredient names in recipes: {len(raw_ingredients)}")
print(f"Matched to product list: {len(matched)}")
print(f"MISSING from product list: {len(unmatched)}")
print()
print("=" * 100)
print(f"{'S.No.':<6s} {'Missing Ingredient':<55s} {'UOM':<8s} {'Used In'}")
print("=" * 100)

for i, (name, data) in enumerate(sorted(unmatched.items(), key=lambda x: x[0].lower()), 1):
    dishes_str = ', '.join(sorted(data['dishes']))
    if len(dishes_str) > 60:
        dishes_str = dishes_str[:57] + '...'
    print(f"{i:<6d} {name:<55s} {data['uom']:<8s} {dishes_str}")

print("=" * 100)
print(f"\nTotal missing: {len(unmatched)} ingredients")
