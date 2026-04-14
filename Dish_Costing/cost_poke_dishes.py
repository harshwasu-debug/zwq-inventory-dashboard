import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

with open(r"E:\Cloud Kitchen\AI Teams\canonical_price_list.json", "r") as f:
    prices_data = json.load(f)

price_map = {}
for item in prices_data["items"]:
    price_map[item["ingredient"].strip().lower()] = {
        "price": item["price_per_unit"], "uom": item["uom"]
    }

# User corrections
price_map["beef tenderloin (india)"] = {"price": 37.50, "uom": "Kg"}
price_map["chicken breast (aurora)"] = {"price": 13.00, "uom": "Kg"}
price_map["chicken thigh boneless"] = {"price": 13.00, "uom": "Kg"}
price_map["chicken wings uncooked"] = {"price": 13.00, "uom": "Kg"}
price_map["beef mince"] = {"price": 43.00, "uom": "Kg"}
price_map["tap water"] = {"price": 0, "uom": "L"}

alias = {
    "cooking oil": "oil vegetable",
    "arwa 500ml": "arwa water",
    "frying oil": "oil vegetable",
    "tortilla flour (12 inch)": "tortilla 12 inch",
    "parmesan cheese, grated": "cheese parmesan grated",
    "jalapeno, sliced": "jalapeno sliced",
    "olives, sliced": "olives black sliced",
    "caesar sauce": "caesar dressing",
}

# Tuna: use Tuna Chunks in Water @ AED 33 / 1.7Kg = 19.41/Kg
price_map["tuna"] = {"price": 19.41, "uom": "Kg"}

# Furikake: SF yields 5.5Kg batch, used as ~5g garnish per bowl
# 1 Piece = 1 portion = 0.005 Kg * 3.65/Kg = ~0.02 AED
# Override: treat "1 Piece" of Furikake as 0.02 AED
piece_overrides = {
    "gyoza wrapper": 0.232,
    "udon noodles": 2.90,
    "egg ramen noodles": 2.95,
    "chicken tenders frozen": 0.045 * 22.0,  # 45g per tender * 22/Kg
    "garnish - house furikake": 0.02,  # ~5g sprinkle from 5.5Kg batch at 3.65/Kg
}

# SF aliases for American-style SFs referenced in wraps
sf_alias = {
    "signature sauce (sf - bulk)": "Sauce - Poke Sauce",
    "tangy buffalo sauce": "buffalo wings sauce",
    "mexican ranch sauce": "Sauce - Poke Sauce",
    "honey chipotle sauce": "Sauce - Spicy Mayo",
    "chili hoisin sauce": "Sauce - Umami Sesame",
}

def gprice(name, uom="Kg", qty=1):
    n = name.strip().lower()
    if uom == "Piece" and n in piece_overrides:
        return piece_overrides[n] * qty
    if n in price_map:
        return price_map[n]["price"] * qty
    if n in alias and alias[n] in price_map:
        return price_map[alias[n]]["price"] * qty
    return None

wb = openpyxl.load_workbook(
    r"C:\Users\harsh\Desktop\Cloud Kitchen\Brands Home\Poke Brands\Poke Recipe.xlsx",
    data_only=True
)

# SF recipes from 'Semi Finished' sheet
ws = wb["Semi Finished"]
sf = {}
for r in range(3, ws.max_row + 1):
    d = ws.cell(row=r, column=2).value
    i = ws.cell(row=r, column=3).value
    q = ws.cell(row=r, column=13).value
    u = ws.cell(row=r, column=14).value
    w = ws.cell(row=r, column=15).value
    if not d or not i:
        continue
    sf.setdefault(d, []).append({
        "i": str(i).strip(), "q": float(q or 0),
        "u": str(u or "Kg"), "w": float(w or 0)
    })

# Compute SF costs iteratively
sf_cost = {}
for _ in range(10):
    changed = False
    for name, ings in sf.items():
        if name in sf_cost:
            continue
        cost = 0; yld = 0; ok = True
        for ig in ings:
            q = ig["q"]; w = ig["w"] / 100
            g = q / (1 - w) if 0 < w < 1 else q
            sf_match = None
            for sn in sf_cost:
                if sn.lower() == ig["i"].lower():
                    sf_match = sn; break
            if sf_match:
                if ig["u"] in ("Kg", "L"):
                    cost += g * sf_cost[sf_match]["ppu"]
                else:
                    cost += g * sf_cost[sf_match]["total"]
            else:
                unresolved = False
                for sn in sf:
                    if sn.lower() == ig["i"].lower() and sn not in sf_cost:
                        unresolved = True; break
                if unresolved:
                    ok = False; break
                p = gprice(ig["i"], ig["u"], g)
                if p is not None:
                    cost += p
            yld += q
        if ok:
            sf_cost[name] = {
                "total": round(cost, 4), "yield": round(yld, 4),
                "ppu": round(cost / yld, 4) if yld > 0 else cost,
            }
            changed = True
    if not changed:
        break

# Finished recipes from 'Sheet4'
ws = wb["Sheet4"]
fin = {}; order = []; last_cat = ""
for r in range(3, ws.max_row + 1):
    cat = ws.cell(row=r, column=1).value
    d = ws.cell(row=r, column=2).value
    i = ws.cell(row=r, column=3).value
    q = ws.cell(row=r, column=13).value
    u = ws.cell(row=r, column=14).value
    w = ws.cell(row=r, column=15).value
    if cat: last_cat = cat
    if not d or not i: continue
    if d not in fin:
        fin[d] = {"cat": last_cat, "ings": []}
        order.append(d)
    if not fin[d]["cat"]: fin[d]["cat"] = last_cat
    fin[d]["ings"].append({
        "i": str(i).strip(), "q": float(q or 0),
        "u": str(u or "Kg"), "w": float(w or 0)
    })

# Also load American SFs for wrap ingredients
try:
    with open(r"E:\Cloud Kitchen\AI Teams\american_dish_costing.json", "r") as f:
        am_data = json.load(f)
    am_sf = am_data.get("sf_costs", {})
except:
    am_sf = {}

# Cost finished dishes - first pass
dish_cost_map = {}
results = []; missing_all = set()
for dn in order:
    data = fin[dn]; cost = 0; miss = []
    for ig in data["ings"]:
        q = ig["q"]; w = ig["w"] / 100
        g = q / (1 - w) if 0 < w < 1 else q
        il = ig["i"].lower()
        # Check piece overrides FIRST (e.g. Furikake garnish)
        if ig["u"] == "Piece" and il in piece_overrides:
            cost += piece_overrides[il] * g
            continue
        # Check Poke SFs
        sf_match = None
        for sn in sf_cost:
            if sn.lower() == il: sf_match = sn; break
        # Check SF alias
        if not sf_match and il in sf_alias:
            target = sf_alias[il]
            for sn in sf_cost:
                if sn.lower() == target.lower(): sf_match = sn; break
        # Check American SFs
        if not sf_match:
            for sn in am_sf:
                if sn.lower() == il:
                    sf_match = sn
                    sf_cost[sn] = am_sf[sn]  # import it
                    break
        if sf_match:
            if ig["u"] in ("Kg", "L"):
                cost += g * sf_cost[sf_match]["ppu"]
            else:
                cost += g * sf_cost[sf_match]["total"]
        else:
            p = gprice(ig["i"], ig["u"], g)
            if p is not None:
                cost += p
            else:
                miss.append(ig["i"]); missing_all.add(ig["i"])
    dish_cost_map[dn.lower().strip()] = cost
    results.append({"dish": dn, "cat": data["cat"], "cost": round(cost, 2), "miss": miss})

# Second pass: resolve combos referencing other finished dishes
for idx, dn in enumerate(order):
    if not results[idx]["miss"]:
        continue
    data = fin[dn]; cost = 0; miss = []
    for ig in data["ings"]:
        q = ig["q"]; w = ig["w"] / 100
        g = q / (1 - w) if 0 < w < 1 else q
        il = ig["i"].lower().strip()
        if ig["u"] == "Piece" and il in piece_overrides:
            cost += piece_overrides[il] * g
            continue
        sf_match = None
        for sn in sf_cost:
            if sn.lower() == il: sf_match = sn; break
        if sf_match:
            if ig["u"] in ("Kg", "L"):
                cost += g * sf_cost[sf_match]["ppu"]
            else:
                cost += g * sf_cost[sf_match]["total"]
        elif il in dish_cost_map:
            cost += dish_cost_map[il] * g
        else:
            p = gprice(ig["i"], ig["u"], g)
            if p is not None:
                cost += p
            else:
                miss.append(ig["i"])
    results[idx]["cost"] = round(cost, 2)
    results[idx]["miss"] = miss
    dish_cost_map[dn.lower().strip()] = cost

missing_all = set()
for r in results:
    for m in r["miss"]:
        missing_all.add(m)

for i, r in enumerate(results, 1):
    m = " [!" + ",".join(set(r["miss"])) + "]" if r["miss"] else ""
    print(f'{i:>3d}. [{r["cat"]:<25s}] {r["dish"]:<45s} AED {r["cost"]:>8.2f}{m}')
print(f"\nTotal: {len(results)} dishes")
if missing_all:
    print(f"Missing: {sorted(missing_all)}")

# Print SF costs
print("\nSF Costs:")
for k, v in sorted(sf_cost.items()):
    print(f"  {k:<40s} total={v['total']:>8.2f} ppu={v['ppu']:>8.2f}")

with open(r"E:\Cloud Kitchen\AI Teams\poke_dish_costing.json", "w") as f:
    json.dump({"dishes": results, "sf_costs": {k: v for k, v in sf_cost.items()}}, f, indent=2)
print("\nSaved to poke_dish_costing.json")
