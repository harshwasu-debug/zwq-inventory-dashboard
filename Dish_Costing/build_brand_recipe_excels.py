import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load source recipe data (Sheet3 = Before Noon master) ─────────────────────
SRC = r'C:\Users\harsh\Desktop\Cloud Kitchen\Brands Home\Breakfast Brands\Breakfast Brands Recipe.xlsx'
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_src = wb_src['Sheet3']

# Parse into: {dish_name: {'category': str, 'rows': [(ingredient, qty, uom, wastage)]}}
recipes = {}
current_cat = ''
for row in ws_src.iter_rows(values_only=True):
    _, dish, ingredient, qty, uom, wastage = row[0], row[1], row[2], row[3], row[4], row[5]
    if not dish:
        continue
    # Skip header row
    if dish == 'Dish' or ingredient == 'Ingredient':
        continue
    # Category header rows have no ingredient
    if dish and not ingredient:
        current_cat = dish
        continue
    # Skip if wastage is not numeric
    try:
        wastage_val = float(wastage) if wastage is not None else 0.0
    except (ValueError, TypeError):
        wastage_val = 0.0
    if dish not in recipes:
        recipes[dish] = {'category': current_cat, 'rows': []}
    recipes[dish]['rows'].append((ingredient, qty, uom, wastage_val))

# ── Dish name mapping: Before Noon name → {brand: brand_dish_name} ────────────
# Structure: BN_name: {'BC': ..., 'TC': ..., 'SC': ...}
# None = dish not on that brand; '' = same meaning, use BN name
DISH_MAP = {
    # TOASTS
    'Truffle Mushroom Scramble': {
        'BC': 'The Truffle Mushroom Scramble',
        'TC': 'Truffle Mushrooms Scramble',
        'SC': 'Truffled Mushroom Scramble',
    },
    'Classic Smashed Avo': {
        'BC': 'Classic Smashed Avo Toast',
        'TC': 'Classic Smash Avo',
        'SC': 'Classic Smashed Avocado',
    },
    'Mediterranean Zaatar Toast': {
        'BC': 'The Mediterranean Zaatar Toast',
        "TC": "Mediterranean Za'atar Toast",
        "SC": "Mediterranean Za'atar Toasts",
    },
    'The Smoked Salmon Toast': {
        'BC': None,
        'TC': 'Smoked Salmon Toast',
        'SC': 'The Smoked Salmon Toasts',
    },
    'Caprese & Pesto Toast': {
        'BC': 'Caprese and Pesto Toast',
        'TC': 'Caprese & Pesto Toasts',
        'SC': None,
    },
    'Spicy Tuna Melt': {
        'BC': 'Spicy Tuna Melt Toast',
        'TC': 'The Spicy Tuna Melt',
        'SC': None,
    },
    # OMELETTES
    'Three Cheese Fold': {
        'BC': 'The Three Cheese Fold',
        'TC': 'Three-Cheese Fold',
        'SC': 'Triple Cheese Fold',
    },
    'The Egg White (Fit)': {
        'BC': 'The Egg Whites (Fit)',
        'TC': 'Egg White (Fit)',
        'SC': 'Egg Whites (Fit)',
    },
    'Spanish Potato Omelet': {
        'BC': 'Classic Spanish Potato Omelet',
        'TC': 'Spanish Potato Omelette',
        'SC': 'The Spanish Potato Omelet',
    },
    'Mexican Omelet': {
        'BC': 'The Mexican Omelet',
        'TC': None,
        'SC': None,
    },
    # EGGS
    'Levantine Shakshuka': {
        'BC': 'Homestyle Levantine Shakshuka',
        'TC': 'The Levantine Shakshuka',
        'SC': 'Classic Levantine Shakshuka',
    },
    'Turkish Eggs (Cilbir)': {
        'BC': 'Turkish Style Eggs (Cilbir)',
        'TC': 'Turkish Egg (Cilbir)',
        'SC': 'The Turkish Eggs (Cilbir)',
    },
    'The Full American Plate': {
        'BC': 'The Full American Breakfast',
        'TC': 'Full American Plate',
        'SC': None,
    },
    'Steak & Eggs': {
        'BC': None,
        'TC': 'Steak and Eggs',
        'SC': 'Steak & Egg',
    },
    'Scrambled Egg & Soldiers': {
        'BC': None,
        'TC': None,
        'SC': None,
    },
    # BAGELS
    'The B.E.C Bagel': {
        'BC': 'The B.E.C. Bagel',
        'TC': 'The BEC Bagel',
        'SC': 'B.E.C. Bagel',
    },
    'Lox & Cream Cheese': {
        'BC': 'Lox & Cream Cheese Bagel',
        'TC': 'The Lox & Cream Cheese',
        'SC': 'Lox and Cream Cheese',
    },
    'The Halloumi Bagel': {
        'BC': 'Halloumi Cheese Bagel',
        'TC': 'Halloumi Bagel',
        'SC': 'Grilled Halloumi Bagel',
    },
    'Avocado & Egg Bagel': {
        'BC': 'The Avocado & Egg Bagel',
        'TC': 'Avocados & Egg Bagel',
        'SC': 'Avocado & Egg Bagels',
    },
    # BUNS & BURGERS
    'Breakfast Bao Trio': {
        'BC': 'The Breakfast Bao Trio',
        'TC': None,
        'SC': None,
    },
    'Crispy Chicken Morning Bao': {
        'BC': 'The Crispy Chicken Morning Bao',
        'TC': None,
        'SC': None,
    },
    'The Fairfax Egg Bun': {
        'BC': 'Fairfax Egg Bun',
        'TC': None,
        'SC': None,
    },
    'The Morning Burger': {
        'BC': None,
        'TC': None,
        'SC': None,
    },
    # BURRITOS & TACOS
    'Ranchero Burrito': {
        'BC': 'Classic Ranchero Burrito',
        'TC': None,
        'SC': 'The Ranchero Burrito',
    },
    'Breakfast Tacos (3 Pcs)': {
        'BC': 'The Breakfast Tacos (3 Pcs)',
        'TC': None,
        'SC': 'Power Breakfast Tacos (3 Pcs)',
    },
    # BOWLS
    'Grilled Chicken Protein Bowl': {
        'BC': 'Classic Grilled Chicken Bowl',
        'TC': 'The Grilled Chicken Protein Bowl',
        'SC': 'Grilled Chicken Power Bowl',
    },
    'Middle Eastern Breakfast Bowl': {
        'BC': 'Classic Middle Eastern Breakfast Bowl',
        'TC': 'The Middle Eastern Breakfast Bowl',
        'SC': 'Middle Eastern Nourish Bowl',
    },
    'Protein Power Bowl': {
        'BC': None,   # BC has Hearty Protein Bowl (different recipe)
        'TC': 'The Protein Power Bowl',
        'SC': 'Power Protein Bowl',
    },
    'Salmon Super Bowl': {
        'BC': None,
        'TC': 'The Salmon Super Bowl',
        'SC': 'Salmon Superfood Bowl',
    },
    # AÇAÍ BOWLS
    'Classic Açaí Bowl': {
        'BC': None,
        'TC': 'Classic Acai Bowl',
        'SC': 'The Classic Açaí Bowl',
    },
    'Tropical Açaí Bowl': {
        'BC': None,
        'TC': 'Tropical Acai Bowl',
        'SC': 'Tropical Açaí Power Bowl',
    },
    'Protein Açaí Bowl': {
        'BC': None,
        'TC': 'Protein Acai Bowl',
        'SC': 'The Protein Açaí Bowl',
    },
    'Date & Pistachio Açaí Bowl': {
        'BC': None,
        'TC': 'Date & Pistachio Acai Bowl',
        'SC': 'Dates & Pistachio Açaí Bowl',
    },
    'Peanut Butter Açaí Bowl': {
        'BC': None,
        'TC': 'Peanut Butter Acai Bowl',
        'SC': 'The Peanut Butter Açaí Bowl',
    },
    'Berry Blast Açaí Bowl': {
        'BC': None,
        'TC': None,
        'SC': 'The Berry Blast Açaí Bowl',
    },
    # PANCAKES & FRENCH TOAST
    'Lotus Biscoff Pancakes': {
        'BC': 'The Lotus Biscoff Pancakes',
        'TC': 'Lotus Biscoff Pancake',
        'SC': None,
    },
    'Nutella & Berry Pancakes': {
        'BC': 'Nutella and Berry Pancakes',
        'TC': 'Nutella & Berries Pancakes',
        'SC': None,
    },
    'Classic French Toast': {
        'BC': 'Homestyle French Toast',
        'TC': 'The Classic French Toast',
        'SC': None,
    },
    'Mini Pancakes (5 Pcs)': {
        'BC': 'Mini Pancake Stack (5 pcs)',
        'TC': None,
        'SC': None,
    },
    # KIDS
    'Scrambled Egg & Soldiers (Kids)': {
        'BC': 'Scrambled Eggs & Soldiers',
        'TC': None,
        'SC': None,
    },
    'Kids French Toast': {
        'BC': 'Kids Classic French Toast',
        'TC': None,
        'SC': None,
    },
    'Nutella Toast': {
        'BC': 'Warm Nutella Toast',
        'TC': None,
        'SC': None,
    },
    'Mini Bagel & Cream Cheese': {
        'BC': 'The Mini Bagel & Cream Cheese',
        'TC': None,
        'SC': None,
    },
    # SIDES
    'Hashbrowns (2 Pcs)': {
        'BC': 'Crispy Hashbrowns (2 Pcs)',
        'TC': 'Hash Browns (2 Pcs)',
        'SC': 'Hash Brown (2 Pcs)',
    },
    'Crispy Beef Bacon': {
        'BC': 'Beef Bacon (3 Strips)',
        'TC': 'The Crispy Beef Bacon',
        'SC': 'Beef Bacon Strips',
    },
    'Chicken Sausage': {
        'BC': 'Chicken Sausage Patties',
        'TC': 'Chicken Sausages',
        'SC': 'Chicken Sausage Links',
    },
    'Grilled Halloumi': {
        'BC': None,
        'TC': 'The Grilled Halloumi',
        'SC': 'Grilled Halloumi Slices',
    },
    'Sliced Avocado': {
        'BC': 'Avocado Slices',
        'TC': 'Sliced Avocados',
        'SC': 'Fresh Sliced Avocado',
    },
    'Fruit Cup': {
        'BC': 'Classic Fruit Cup',
        'TC': 'Fresh Fruit Cup',
        'SC': 'Seasonal Fruit Cup',
    },
    # SAUCES
    'Truffle Mayo': {
        'BC': 'House Truffle Mayo',
        'TC': 'Truffled Mayo',
        'SC': 'Truffle Aioli',
    },
    'Sriracha Mayo': {
        'BC': 'Classic Sriracha Mayo',
        'TC': 'Sriracha Chili Mayo',
        'SC': 'Sriracha Sauce',
    },
    'Chipotle Sauce': {
        'BC': 'Smoky Chipotle Sauce',
        'TC': 'Chipotle Mayo',
        'SC': 'Chipotle Aioli',
    },
    'Guacamole': {
        'BC': 'Homestyle Guacamole',
        'TC': 'Fresh Guacamole',
        'SC': 'Classic Guacamole',
    },
    'Lotus Biscoff Sauce': {
        'BC': 'Lotus Biscoff Dip',
        'TC': 'Biscoff Sauce',
        'SC': 'Lotus Biscoff Drizzle',
    },
    'Maple Syrup': {
        'BC': 'Warm Maple Syrup',
        'TC': 'Pure Maple Syrup',
        'SC': 'Raw Maple Syrup',
    },
    # OVERNIGHT OATS
    'The Original': {
        'BC': 'Vanilla Classic Oats',
        'TC': 'Original Oats',
        'SC': 'Pure Oats',
    },
    'Tropical Sunrise': {
        'BC': None,
        'TC': None,
        'SC': 'The Tropical Sunrise',
    },
    'Berry Blast': {
        'BC': 'Classic Berry Blast Oats',
        'TC': 'The Berry Blast',
        'SC': 'Berry Power Oats',
    },
    'Banana & Peanut Butter': {
        'BC': 'Banana Peanut Butter Overnight Oats',
        'TC': 'Banana and Peanut Butter',
        'SC': 'Banana & Peanut Butter Oats',
    },
    'Mango & Berry': {
        'BC': None,
        'TC': 'Mango & Berries',
        'SC': 'Mango & Berry Oats',
    },
    'Apple & Date': {
        'BC': 'Apple and Date Oats',
        'TC': 'Apple & Dates',
        'SC': 'Apple & Date Oats',
    },
    'Tropical Berry': {
        'BC': None,
        'TC': None,
        'SC': 'Tropical Berry Oats',
    },
    'Classic Banana Split': {
        'BC': 'The Banana Split Oats',
        'TC': None,
        'SC': None,
    },
}

# Dishes unique to Breakfast Counter (no BN equivalent)
BC_UNIQUE = {
    'Hearty Protein Bowl': {
        'category': 'PROTEIN BOWLS',
        'note': 'Recipe TBD — see Deliverect description: Scrambled eggs, grilled chicken, crispy beef bacon, veggies and cheese over seasoned rice with Classic Sriracha Mayo.',
        'rows': []
    }
}

# ── Styles ─────────────────────────────────────────────────────────────────────
def make_styles(brand_color):
    hdr_font  = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    hdr_fill  = PatternFill(start_color=brand_color, end_color=brand_color, fill_type='solid')
    cat_font  = Font(bold=True, color='FFFFFF', size=10, name='Arial')
    cat_fill  = PatternFill(start_color='555555', end_color='555555', fill_type='solid')
    ing_font  = Font(name='Arial', size=10)
    pkg_fill  = PatternFill(start_color='EAF2E3', end_color='EAF2E3', fill_type='solid')
    tbd_fill  = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    thin      = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    return hdr_font, hdr_fill, cat_font, cat_fill, ing_font, pkg_fill, tbd_fill, thin

BRAND_COLORS = {
    'BN': 'CF6228',   # Before Noon — Terracotta
    'BC': '5B4A3F',   # Breakfast Counter — Warm Brown
    'TC': '2F5496',   # Toast & Co — Navy Blue
    'SC': 'E07B39',   # Sunrise & Co — Sunrise Orange
}

BRAND_NAMES = {
    'BN': 'Before Noon',
    'BC': 'Breakfast Counter',
    'TC': 'Toast & Co',
    'SC': 'Sunrise & Co',
}

OUTPUT_DIR = r'C:\Users\harsh\Desktop\Cloud Kitchen\Brands Home\Breakfast Brands\WIP files'

# ── Packaging keyword detection ────────────────────────────────────────────────
PKG_KEYWORDS = {'tray', 'lid', 'bag', 'cup', 'cutlery', 'box', 'container',
                'wrap', 'fork', 'spoon', 'napkin', 'straw', 'paper', 'foil',
                'sleeve', 'label', 'seal', 'sachet', 'bottle'}

def is_packaging(ingredient_name):
    lower = ingredient_name.lower()
    return any(kw in lower for kw in PKG_KEYWORDS)

# ── Build one workbook per brand ───────────────────────────────────────────────
def build_brand_excel(brand_code):
    color = BRAND_COLORS[brand_code]
    brand_name = BRAND_NAMES[brand_code]
    hdr_font, hdr_fill, cat_font, cat_fill, ing_font, pkg_fill, tbd_fill, thin = make_styles(color)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Full Recipes ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Recipes'

    # Header row
    headers = ['#', 'Category', 'Dish Name', 'Ingredient', 'Qty (EP)', 'UOM', 'Wastage %', 'Type']
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font  = hdr_font
        c.fill  = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # Build ordered dish list for this brand
    brand_dishes = []  # (display_name, category, bn_source_name, rows)

    seen_cats = set()
    for bn_name, data in recipes.items():
        if brand_code == 'BN':
            display_name = bn_name
            cat = data['category']
            rows = data['rows']
        else:
            mapping = DISH_MAP.get(bn_name, {})
            display_name = mapping.get(brand_code)
            if display_name is None:
                continue  # dish not on this brand
            cat = data['category']
            rows = data['rows']

        brand_dishes.append((display_name, cat, bn_name, rows))

    # Add brand-unique dishes
    if brand_code == 'BC':
        for unique_name, unique_data in BC_UNIQUE.items():
            brand_dishes.append((unique_name, unique_data['category'], None, unique_data['rows']))

    # Write rows
    row_num = 2
    dish_num = 0
    current_cat = None

    for display_name, cat, bn_name, ing_rows in brand_dishes:
        # Category separator row
        if cat != current_cat:
            current_cat = cat
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            c = ws.cell(row=row_num, column=1, value=f'◆  {cat}')
            c.font  = cat_font
            c.fill  = cat_fill
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        dish_num += 1

        if not ing_rows:
            # TBD dish — single placeholder row
            vals = [dish_num, cat, display_name, '— Recipe not yet defined —', '', '', '', '']
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=row_num, column=j, value=v)
                c.border = thin
                c.fill   = tbd_fill
                if j == 3:
                    c.font = Font(bold=True, name='Arial', size=10)
                if j == 4:
                    c.font = Font(italic=True, name='Arial', size=10, color='7F7F7F')
            row_num += 1
            continue

        # Ingredient rows
        first = True
        for ingredient, qty, uom, wastage in ing_rows:
            ing_type = 'Packaging' if is_packaging(ingredient) else 'Food'
            d_num_val = dish_num if first else ''
            cat_val   = cat if first else ''
            name_val  = display_name if first else ''

            vals = [d_num_val, cat_val, name_val, ingredient, qty, uom,
                    f'{int(wastage*100)}%' if wastage else '0%', ing_type]  # noqa
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=row_num, column=j, value=v)
                c.border = thin
                c.font   = ing_font
                if j == 3 and first:
                    c.font = Font(bold=True, name='Arial', size=10)
                if j in (5,):
                    c.alignment = Alignment(horizontal='right')
                if ing_type == 'Packaging':
                    c.fill = pkg_fill
            row_num += 1
            first = False

        # Blank separator between dishes
        row_num += 1

    # Column widths
    widths = {'A': 5, 'B': 20, 'C': 36, 'D': 40, 'E': 10, 'F': 8, 'G': 10, 'H': 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    s_headers = ['#', 'Category', 'Dish Name', 'Ingredients', 'Packaging Items', 'Total Lines']
    for j, h in enumerate(s_headers, 1):
        c = ws2.cell(row=1, column=j, value=h)
        c.font  = hdr_font
        c.fill  = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin
    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = 'A2'

    for i, (display_name, cat, bn_name, ing_rows) in enumerate(brand_dishes, 1):
        food_lines = sum(1 for r in ing_rows if not is_packaging(r[0]))
        pkg_lines  = sum(1 for r in ing_rows if is_packaging(r[0]))
        total_lines = len(ing_rows)
        row_data = [i, cat, display_name, food_lines, pkg_lines, total_lines]
        for j, v in enumerate(row_data, 1):
            c = ws2.cell(row=i+1, column=j, value=v)
            c.border = thin
            c.font   = Font(name='Arial', size=10)
            if j in (4, 5, 6):
                c.alignment = Alignment(horizontal='center')
            if total_lines == 0:
                c.fill = tbd_fill

    for col, w in [('A', 5), ('B', 22), ('C', 40), ('D', 14), ('E', 14), ('F', 13)]:
        ws2.column_dimensions[col].width = w

    # Save
    out_path = f'{OUTPUT_DIR}\\{brand_name.replace(" ", "_")}_Recipes.xlsx'
    wb.save(out_path)
    print(f'Saved: {out_path}  ({len(brand_dishes)} dishes)')
    return len(brand_dishes)

# ── Build all 4 ───────────────────────────────────────────────────────────────
print('Building brand recipe Excel files...')
print()
totals = {}
for code in ['BN', 'BC', 'TC', 'SC']:
    n = build_brand_excel(code)
    totals[BRAND_NAMES[code]] = n

print()
print('Done.')
for brand, n in totals.items():
    print(f'  {brand:<22s}: {n} dishes')
