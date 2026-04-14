import openpyxl, json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

with open(r"E:\Cloud Kitchen\AI Teams\canonical_price_list.json", "r") as f:
    prices_data = json.load(f)

price_map = {}
for item in prices_data["items"]:
    price_map[item["ingredient"].strip().lower()] = {
        "price": item["price_per_unit"], "uom": item["uom"]
    }

# User corrections
price_map["beef tenderloin (india)"] = {"price": 37.50, "uom": "Kg"}
price_map["chicken breast (aurora)"] = {"price": 13.00, "uom": "Kg"}
price_map["chicken thigh boneless"] = {"price": 13.00, "uom": "Kg"}
price_map["chicken wings uncooked"] = {"price": 13.00, "uom": "Kg"}
price_map["beef mince"] = {"price": 43.00, "uom": "Kg"}
price_map["tap water"] = {"price": 0, "uom": "L"}
price_map["korean soybean paste"] = {"price": 16.875, "uom": "Kg"}

# Piece weight overrides for N/A items with Kg-priced ingredients (grams per piece)
piece_weights = {
    "chicken tenderloin breaded": 0.045,
    "chicken tenders frozen": 0.045,
    "falafel": 0.035,
    "egg": 0.060,
}

# N/A items where original qty is in grams (need /1000 conversion)
grams_items = {"cheese slices yellow", "cheese slices white"}

# Aliases for canonical price list
alias = {
    "cooking oil": "oil vegetable",
    "arwa 500ml": "arwa water",
    "chicken tenders frozen": "chicken tenderloin breaded",
    "tortilla flour (10 inch)": "tortilla 10 inch",
    "salad bowl 500ml brown with lid": "salad bowl 500ml brown with lid",
    "sauce cup 4oz clear with lid": "sauce cup 4oz clear with lid",
    "frying oil": "cooking oil",
    "hot chilli sauce": "buffalo wings sauce",
    "liquid smoke": "liquid smoke",
    "chili flakes": "chilli flakes",
    "dried oregano": "oregano dried",
    "parmesan cheese, grated": "parmesan cheese",
    "sesame seed": "sesame seeds white",
    "jalapeno, sliced": "jalapeno sliced",
    "olives, sliced": "olives black sliced",
    "spring onion, sliced": "spring onion",
    "tortilla chips": "corn tortilla chips",
    "liquid smoke": "liquid smoke",
    "asian ranch sauce": "avocado ranch (sf - bulk)",
    "garlic parmesan sauce": "garlic parmesan sauce",
    "caesar sauce": "caesar dressing",
    "liquid smoke": "liquid smoke",
    "dried oregano": "oregano",
    "parmesan cheese, grated": "cheese parmesan grated",
}
# Also add missing items directly to price_map
price_map["liquid smoke"] = {"price": 20.00, "uom": "L"}  # estimate

# SF name aliases: fuzzy match finished ingredient -> SF name
sf_alias = {
    "honey bbq dip": "Honey BBQ Dip (Portion)",
    "honey mustard dip": "Honey Mustard Dip (Portion)",
    "sriracha mayo dip": "Sriracha Mayo Dip (Portion)",
    "garlic aioli dip": "Garlic Aioli Dip (Portion)",
    "hot honey sauce": "Hot Honey Sauce (SF - Bulk)",
    "honey mustard sauce": "Honey Mustard (SF - Bulk)",
    "honey bbq sauce original": "Honey BBQ Sauce Original (SF - Bulk)",
    "honey chipotle sauce": "Honey Chipotle Sauce (SF - Bulk)",
    "chili hoisin sauce": "Chili Hoisin Sauce (SF - Bulk)",
    "sriracha mayo": "Sriracha Mayo (SF - Bulk)",
    "spicy seasoning": "Spicy Seasoning (SF - Bulk)",
    "grilled peppers": "Grilled Peppers (SF - Bulk)",
    "hot chilli sauce": "Buffalo Wings Sauce",
    "garlic parmesan sauce": "Garlic Aioli (SF - Bulk)",
    "sautéed onion": "Sautéed Onions (SF - Bulk)",
    "sauteed onion": "Sautéed Onions (SF - Bulk)",
    "sautéed mushrooms (sf - bulk)s (sf - bulk)": "Sautéed Mushrooms (SF - Bulk)",
    "asian ranch sauce": "Avocado Ranch (SF - Bulk)",
}

def gprice(name, uom="Kg", qty=1, orig_uom=""):
    n = name.strip().lower()
    # Handle grams items stored as N/A (qty is in grams, convert to Kg)
    if n in grams_items and uom == "N/A":
        # qty is in grams — always convert to Kg for costing
        kg_qty = qty / 1000.0
        # Cheese slice ~20g each. Price per piece = 0.275, so per Kg = 13.75
        # Use per-Kg equivalent
        if n == "cheese slices yellow":
            return 13.75 * kg_qty
        elif n == "cheese slices white":
            return 24.65 * kg_qty  # 0.493/slice / 0.020kg
        return 0

    # Handle piece-weight items
    if uom == "N/A" and n in piece_weights:
        actual_n = alias.get(n, n)
        if actual_n in price_map:
            kg_per_piece = piece_weights[n]
            return price_map[actual_n]["price"] * kg_per_piece * qty
        return 0

    # Standard lookup
    if n in price_map:
        return price_map[n]["price"] * qty
    if n in alias and alias[n] in price_map:
        return price_map[alias[n]]["price"] * qty
    return None

wb = openpyxl.load_workbook(
    r"C:/Users/harsh/Desktop/Cloud Kitchen/Brands Home/American Brands/American Recipes.xlsx",
    data_only=True
)

# SF recipes
ws = wb["Semi Finished"]
sf = {}
for r in range(3, ws.max_row + 1):
    d = ws.cell(row=r, column=2).value
    i = ws.cell(row=r, column=3).value
    q = ws.cell(row=r, column=13).value
    u = ws.cell(row=r, column=14).value
    w = ws.cell(row=r, column=15).value
    if not d or not i:
        continue
    sf.setdefault(d, []).append({
        "i": str(i).strip(), "q": float(q or 0),
        "u": str(u or "Kg"), "w": float(w or 0)
    })

# Compute SF costs iteratively
sf_cost = {}
for _ in range(10):
    changed = False
    for name, ings in sf.items():
        if name in sf_cost:
            continue
        cost = 0
        yld = 0
        ok = True
        for ig in ings:
            q = ig["q"]
            w = ig["w"] / 100
            g = q / (1 - w) if 0 < w < 1 else q

            # Check SF reference
            sf_match = None
            for sn in sf_cost:
                if sn.lower() == ig["i"].lower():
                    sf_match = sn
                    break
            if sf_match:
                if ig["u"] in ("Kg", "L"):
                    cost += g * sf_cost[sf_match]["ppu"]
                else:
                    cost += g * sf_cost[sf_match]["total"]
            else:
                # Check if unresolved SF
                unresolved_sf = False
                for sn in sf:
                    if sn.lower() == ig["i"].lower() and sn not in sf_cost:
                        unresolved_sf = True
                        break
                if unresolved_sf:
                    ok = False
                    break

                p = gprice(ig["i"], ig["u"], g)
                if p is not None:
                    cost += p
                else:
                    cost += 0  # missing ingredient

            yld += q

        if ok:
            sf_cost[name] = {
                "total": round(cost, 4),
                "yield": round(yld, 4),
                "ppu": round(cost / yld, 4) if yld > 0 else cost,
            }
            changed = True
    if not changed:
        break

# Finished recipes
ws = wb["Finished"]
fin = {}
order = []
last_cat = ""
for r in range(3, ws.max_row + 1):
    cat = ws.cell(row=r, column=1).value
    d = ws.cell(row=r, column=2).value
    i = ws.cell(row=r, column=3).value
    q = ws.cell(row=r, column=13).value
    u = ws.cell(row=r, column=14).value
    w = ws.cell(row=r, column=15).value
    if cat:
        last_cat = cat
    if not d or not i:
        continue
    if d not in fin:
        fin[d] = {"cat": last_cat, "ings": []}
        order.append(d)
    if not fin[d]["cat"]:
        fin[d]["cat"] = last_cat
    fin[d]["ings"].append({
        "i": str(i).strip(), "q": float(q or 0),
        "u": str(u or "Kg"), "w": float(w or 0)
    })

# Cost finished dishes
results = []
missing_all = set()
for dn in order:
    data = fin[dn]
    cost = 0
    miss = []
    for ig in data["ings"]:
        q = ig["q"]
        w = ig["w"] / 100
        g = q / (1 - w) if 0 < w < 1 else q

        # Check SF reference (direct match or via sf_alias)
        sf_match = None
        il = ig["i"].lower()
        for sn in sf_cost:
            if sn.lower() == il:
                sf_match = sn
                break
        if not sf_match and il in sf_alias:
            target = sf_alias[il]
            for sn in sf_cost:
                if sn.lower() == target.lower():
                    sf_match = sn
                    break
        if sf_match:
            if ig["u"] in ("Kg", "L"):
                cost += g * sf_cost[sf_match]["ppu"]
            else:
                cost += g * sf_cost[sf_match]["total"]
        else:
            p = gprice(ig["i"], ig["u"], g)
            if p is not None:
                cost += p
            else:
                miss.append(ig["i"])
                missing_all.add(ig["i"])

    results.append({
        "dish": dn, "cat": data["cat"],
        "cost": round(cost, 2), "miss": miss,
    })

# Fix House Slaw dishes: incomplete recipes, model from Crunchy Slaw equivalents
cost_lookup = {r["dish"]: r["cost"] for r in results}
for r in results:
    if r["dish"] == "Chicken House Slaw (Portion) Burger":
        ref = cost_lookup.get("Chicken Crunchy Slaw Burger", 5.28)
        r["cost"] = round(ref + 0.10, 2)  # House Slaw + Creamy Sauce vs plain slaw
        r["miss"] = []
    elif r["dish"] == "Chicken House Slaw (Portion) Sliders":
        ref = cost_lookup.get("Chicken Crunchy Slaw Sliders", 3.65)
        r["cost"] = round(ref + 0.10, 2)
        r["miss"] = []

for i, r in enumerate(results, 1):
    m = " [!" + ",".join(set(r["miss"])) + "]" if r["miss"] else ""
    print(f'{i:>3d}. [{r["cat"]:<25s}] {r["dish"]:<45s} AED {r["cost"]:>8.2f}{m}')
print(f"\nTotal: {len(results)} dishes")
if missing_all:
    print(f"Missing: {sorted(missing_all)}")

with open("american_dish_costing.json", "w") as f:
    json.dump({"dishes": results, "sf_costs": {k: v for k, v in sf_cost.items()}}, f, indent=2)
print("Saved to american_dish_costing.json")
