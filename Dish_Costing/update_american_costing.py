from openpyxl import load_workbook
from openpyxl.styles import Font, numbers

wb = load_workbook(r'E:\Cloud Kitchen\AI Teams\Dish_Costing\Complete_Dish_Costing_Final.xlsx')
ws = wb['American Menu']

# --- 1. Build row index ---
row_map = {}
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=3).value
    if name:
        row_map[name.strip().lower()] = r

# --- 2. Sell price updates (39 mismatches - base menu prices) ---
sell_updates = {
    'classic beef sliders': 64, 'beef slaw sliders': 66, 'buffalo beef sliders': 64,
    'texas beef sliders': 64, 'mushroom truffle beef sliders': 66,
    'beef double cheese sliders': 65, 'double beef sliders': 78,
    'bacon beef sliders': 68, 'bbq crispy beef sliders': 69,
    'mexican crunch beef sliders': 65, 'grilled classic chicken sliders': 51,
    'crispy chicken sliders': 55, 'chicken truffle sliders': 56,
    'honey bbq chicken sliders': 55, 'chicken crunchy slaw sliders': 55,
    'chicken dynamite sliders': 55, 'chicken honey mustard sliders': 55,
    'chicken avocado sliders': 57, 'chicken buffalo sliders': 56,
    'mexican crunch chicken sliders': 56, 'hot honey sliders': 55,
    'asian ranch sliders': 55, 'italian chicken sliders': 54,
    'chili chicken sliders': 56,
    'classic grilled halloumi slider': 52, 'buffalo grilled halloumi slider': 53,
    'bbq grilled halloumi slider': 54, 'avocado grilled halloumi slider': 55,
    'classic falafel slider': 42, 'buffalo falafel slider': 44,
    'bbq falafel slider': 45, 'honey mustard falafel slider': 44,
    'avocado falafel slider': 45,
    'beef sliders meal for 1': 72, 'beef sliders meal for 2': 128,
    'beef sliders meal for 4': 194,
    'chicken sliders meal for 1': 62, 'chicken sliders meal for 2': 108,
    'chicken sliders meal for 4': 172,
}

updated_sell = 0
for name, price in sell_updates.items():
    if name in row_map:
        r = row_map[name]
        ws.cell(row=r, column=5).value = price
        updated_sell += 1
print(f"Updated {updated_sell} sell prices")

# --- 3. Fix costing bugs ---
# Wings Meal for 1: cost should be 9.74 (half of 19.48)
if 'wings meal for 1' in row_map:
    r = row_map['wings meal for 1']
    ws.cell(row=r, column=4).value = 9.74
    print("Fixed Wings Meal for 1 cost: 19.48 -> 9.74")

# Chicken Sliders Meal for 1: cost should be 7.45 (half of 15.05)
# Note: sell price also updated above to 62
if 'chicken sliders meal for 1' in row_map:
    r = row_map['chicken sliders meal for 1']
    ws.cell(row=r, column=4).value = 7.45
    print("Fixed Chicken Sliders Meal for 1 cost: 15.05 -> 7.45")

# --- 4. Add missing dishes ---
# Find the last row
last_row = ws.max_row

# Wings Meal for 1 also needs sell price
if 'wings meal for 1' in row_map:
    r = row_map['wings meal for 1']
    current_sell = ws.cell(row=r, column=5).value
    if current_sell is None:
        # No sell price in base menu for wings meal for 1 either - but pattern suggests ~62-72
        # Not in user's base menu, so leave as is
        pass

# Add missing dishes at the end
new_dishes = [
    # (S.No., Category, Dish Name, Cost, Sell Price)
    ('Beef Burger', 'Classic Beef Burger', 11.84, 66),
    ('Beef Burger', 'Crunchy BBQ Burger', 14.38, 72),
    ('Sliders', 'Crunchy BBQ Sliders', 13.15, 69),
    ('Veggie Sliders', 'Honey Mustard Grilled Halloumi Slider', 8.49, 53),
    ('Sides & Dips', 'Veggie Sticks', None, 19),
]

# Get last S.No.
last_sno = 0
for r in range(2, last_row + 1):
    v = ws.cell(row=r, column=1).value
    if isinstance(v, (int, float)) and v > last_sno:
        last_sno = int(v)

added = 0
for cat, name, cost, sell in new_dishes:
    # Check if already exists
    if name.strip().lower() in row_map:
        print(f"  Skipping {name} - already exists")
        continue
    last_sno += 1
    last_row += 1
    ws.cell(row=last_row, column=1).value = last_sno
    ws.cell(row=last_row, column=2).value = cat
    ws.cell(row=last_row, column=3).value = name
    if cost is not None:
        ws.cell(row=last_row, column=4).value = cost
    ws.cell(row=last_row, column=5).value = sell
    # Add margin formula (col G) = IFERROR(((E-MIN(E/2,30)-4)*0.7)-D,0)
    ws.cell(row=last_row, column=7).value = f'=IFERROR(((E{last_row}-MIN(E{last_row}/2,30)-4)*0.7)-$D{last_row},0)'
    # Margin% (col H)
    ws.cell(row=last_row, column=8).value = f'=G{last_row}/E{last_row}'
    # Ideal Price (col J)
    ws.cell(row=last_row, column=10).value = f'=ROUND(IF(D{last_row}<=3.2,28+10*(D{last_row}),(23.8+D{last_row})/0.45),0)'
    # Final Price (col L)
    ws.cell(row=last_row, column=12).value = f'=MAX(J{last_row},E{last_row})'
    added += 1
    print(f"  Added {name} (Cost: {cost}, Sell: {sell})")

print(f"\nAdded {added} new dishes")

# --- 5. Save ---
output = r'E:\Cloud Kitchen\AI Teams\Dish_Costing\Complete_Dish_Costing_Final.xlsx'
wb.save(output)
print(f"\nSaved to {output}")
