from openpyxl import load_workbook

# Parse the price list - use higher value for duplicates
price_text = """Classic Beef Burger	69
Crunchy BBQ Burger	73
Beef Slaw Burger	71
Beef Burger Quesadilla	71
Buffalo Beef Burger	71
Texas Beef Burger	73
Mushroom Truffle Beef Burger	75
Beef Double Cheese Burger	72
Double Beef Burger	85
Bacon Beef Burger	72
BBQ Crispy Beef Burger	71
Mexican Crunch Beef Burger	71
Grilled Classic Chicken Burger	54
Crispy Chicken Burger	56
Chicken Truffle Burger	56
Honey BBQ Chicken Burger	56
Chicken Crunchy Slaw Burger	56
Chicken Dynamite Burger	56
Chicken Honey Mustard Burger	56
Chicken Avocado Burger	58
Chicken Buffalo Burger	58
Mexican Crunch Chicken Burger	57
Hot Honey Burger	58
Asian Ranch Burger	57
Italian Chicken Burger	58
Chili Chicken Burger	57
Classic Beef Sliders	69
Crunchy BBQ Sliders	74
Beef Slaw Sliders	71
Buffalo Beef Sliders	71
Texas Beef Sliders	71
Mushroom Truffle Beef Sliders	73
Beef Double Cheese Sliders	73
Double Beef Sliders	85
Bacon Beef Sliders	76
BBQ Crispy Beef Sliders	75
Mexican Crunch Beef Sliders	71
Grilled Classic Chicken Sliders	58
Crispy Chicken Sliders	59
Chicken Truffle Sliders	59
Honey BBQ Chicken Sliders	59
Chicken Crunchy Slaw Sliders	59
Chicken Dynamite Sliders	59
Chicken Honey Mustard Sliders	59
Chicken Avocado Sliders	62
Chicken Buffalo Sliders	62
Mexican Crunch Chicken Sliders	62
Hot Honey Sliders	62
Asian Ranch Sliders	62
Italian Chicken Sliders	62
Chili Chicken Sliders	62
Classic Grilled Halloumi Slider	58
Buffalo Grilled Halloumi Slider	58
BBQ Grilled Halloumi Slider	58
Honey Mustard Grilled Halloumi Slider	58
Avocado Grilled Halloumi Slider	59
Classic Falafel Slider	46
Buffalo Falafel Slider	48
BBQ Falafel Slider	48
Honey Mustard Falafel Slider	48
Avocado Falafel Slider	48
Chicken Wings	64
Smokey BBQ Wings	62
Tangy Buffalo Wings	62
Hot Honey Wings	62
Sesame & Sweet Wings	63
Chili Hoisin Wings	62
Garlic Parmesan Wings	62
Honey Mustard Wings	61
Dynamite Wings	61
Chicken Tenders	49
Smokey BBQ Tenders	54
Tangy Buffalo Tenders	52
Hot Honey Tenders	52
Sesame & Sweet Tenders	52
Chili Hoisin Tenders	52
Garlic Parmesan Tenders	52
Honey Mustard Tenders	51
Dynamite Tenders	51
Chicken Wraps	62
Chicken Caesar Wrap	55
BBQ Chicken Wrap	57
Buffalo Chicken Wrap	57
Hot Honey Chicken Wrap	56
Asian Chicken Wrap	56
Fries	21
Spicy Fries	23
Dynamite Fries	28
Cheesy Fries	26
Tex Mex Fries	30
Onion Rings	25
Curly Fries	24
Mozzarella Sticks	29
Dynamite Shrimp	48
Veggie Sticks	19
Nachos	35
Loaded Nachos	52
House Slaw	29
Garlic Aioli Dip	8
Sriracha Mayo Dip	8
Honey BBQ Dip	8
Dynamite Sauce Dip	8
Truffle Mayo Dip	8
Honey Mustard Dip	8
Mexican Ranch Dip	8
Cheese Dip	8
Avocado Ranch Dip	8
Wings Meal for 1	78
Wings Meal for 2	132
Wings Meal for 4	196
Beef Sliders Meal for 1	77
Beef Sliders Meal for 2	126
Beef Sliders Meals for 4	186
Chicken Sliders Meal for 1	70
Chicken Sliders Meal for 2	115
Chicken Sliders Meal for 4	172
Chicken Burger Combo Meal for 1	71
Chicken Burger Combo Meal for 2	118
Chicken Burger Combo Meal for 4	178
Beef Burger Combo Meal for 1	80
Beef Burger Combo Meal for 2	132
Beef Burger Combo Meal for 4	192
Mineral Water	14
Coca Cola	9
Coca Cola Zero	10
Sprite	9
Sprite Zero	10
Fanta	9
Honey Chilli Potatoes	36
Honey Chilli Shrimp	49"""

# Build price map (higher value for duplicates)
price_map = {}
for line in price_text.strip().split('\n'):
    parts = line.rsplit('\t', 1)
    name = parts[0].strip()
    price = int(parts[1].strip())
    if name in price_map:
        price_map[name] = max(price_map[name], price)
    else:
        price_map[name] = price

# Also add aliases for items in costing sheet with different names
aliases = {
    'Beef Sliders Meal for 4': 'Beef Sliders Meals for 4',  # menu has "Meals"
    'Chicken Burger': 'Chicken Burgers',
}

# Load costing sheet
wb = load_workbook(r'E:\Cloud Kitchen\AI Teams\Dish_Costing\American_Menu_Costing.xlsx')
ws = wb['American Menu']

updated = 0
not_found = []
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=3).value
    if not name:
        continue
    name = name.strip()

    # Try exact match
    price = price_map.get(name)

    # Try alias
    if price is None and name in aliases:
        price = price_map.get(aliases[name])

    if price is not None:
        old = ws.cell(row=r, column=5).value
        ws.cell(row=r, column=5).value = price
        if old != price:
            print(f"  Row {r}: {name:45s} | {old} -> {price}")
            updated += 1
    else:
        current = ws.cell(row=r, column=5).value
        not_found.append((r, name, current))

print(f"\nUpdated {updated} sell prices")
if not_found:
    print(f"\nNo price found for {len(not_found)} items (keeping current values):")
    for r, name, current in not_found:
        print(f"  Row {r}: {name} (current: {current})")

wb.save(r'E:\Cloud Kitchen\AI Teams\Dish_Costing\American_Menu_Costing.xlsx')
print("\nSaved.")
